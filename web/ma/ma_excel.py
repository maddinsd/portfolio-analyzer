"""Goldman-quality 8-tab M&A Excel workbook builder.

Tabs: Cover | Assumptions | Transaction | Acquirer | Target | Pro Forma | Accretion | Sensitivity
Color palette: navy headers, blue inputs, green accretive, red dilutive.
"""
from __future__ import annotations

import math
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from ma.ma_fetcher import MACompanyData

# ── Design constants ───────────────────────────────────────────────────────────
_NAVY   = "003366"
_DARK_B = "1F4E79"
_MID_B  = "2E75B6"
_INPUT  = "EBF3FB"
_GRN    = "00B050"
_RED    = "FF0000"
_GRN_BG = "C6EFCE"
_RED_BG = "FFC7CE"
_YLW_BG = "FFEB9C"
_HDR_BG = "D9E1F2"
_WHITE  = "FFFFFF"
_LGRAY  = "F2F2F2"
_BORDER_COLOR = "BFBFBF"

_thin = Side(style="thin", color=_BORDER_COLOR)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _align(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _hdr(ws, row, col, end_col, text: str, bg=_DARK_B, fg=_WHITE, size=10, bold=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _font(bold=bold, color=fg, size=size)
    cell.fill      = _fill(bg)
    cell.alignment = _align("center")
    cell.border    = _BORDER
    if end_col > col:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=end_col
        )


def _label(ws, row, col, text, bold=False, italic=False, indent=0):
    cell = ws.cell(row=row, column=col, value=(" " * indent * 3) + str(text) if indent else text)
    cell.font      = _font(bold=bold, italic=italic)
    cell.alignment = _align("left")
    cell.border    = _BORDER


def _val(ws, row, col, value, fmt="number", color="000000", bold=False, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, color=color)
    cell.alignment = _align("right")
    cell.border    = _BORDER
    if bg:
        cell.fill = _fill(bg)
    if fmt == "dollar":
        cell.number_format = '$#,##0.0;($#,##0.0);"-"'
    elif fmt == "dollar0":
        cell.number_format = '$#,##0;($#,##0);"-"'
    elif fmt == "eps":
        cell.number_format = '$#,##0.00;($#,##0.00)'
    elif fmt == "pct":
        cell.number_format = "0.0%"
    elif fmt == "pct_plain":
        cell.number_format = "0.0%"
    elif fmt == "mult":
        cell.number_format = '0.0"x"'
    elif fmt == "mult2":
        cell.number_format = '0.00"x"'
    elif fmt == "number":
        cell.number_format = "#,##0.0"
    elif fmt == "number0":
        cell.number_format = "#,##0"
    elif fmt == "acct_pct":
        cell.number_format = '+0.0%;-0.0%;"-"'


def _input(ws, row, col, value, fmt="dollar"):
    _val(ws, row, col, value, fmt=fmt, bg=_INPUT)


def _divider(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = _fill(_LGRAY)
        cell.border = _BORDER


def _set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Tab 1: Cover ──────────────────────────────────────────────────────────────

def _build_cover(ws, acq: MACompanyData, tgt: MACompanyData,
                 tx: dict, pf: dict, syn: dict):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 2

    # Title banner
    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 12

    _hdr(ws, 2, 2, 5,
         f"MERGER CONSEQUENCES ANALYSIS",
         bg=_NAVY, fg=_WHITE, size=16, bold=True)
    _hdr(ws, 3, 2, 5,
         f"{acq.lbo.company_name} / {tgt.lbo.company_name}",
         bg=_DARK_B, fg=_WHITE, size=13, bold=True)

    r = 5
    # ── Deal summary block ────────────────────────────────────────────────────
    _hdr(ws, r, 2, 5, "TRANSACTION SUMMARY", bg=_DARK_B, fg=_WHITE, size=10)
    r += 1

    summary_rows = [
        ("Acquirer",            acq.lbo.company_name,                 "",                     ""),
        ("Target",              tgt.lbo.company_name,                 "",                     ""),
        ("Offer Price / Share", f"${tx['offer_price']:.2f}",          "Premium to Unaffected", f"{tx['offer_premium_pct']:.0f}%"),
        ("Total Equity Value",  f"${tx['total_equity_value']:,.0f}M", "Total Enterprise Value",f"${tx['total_ev']:,.0f}M"),
        ("Implied EV/EBITDA",   f"{tx['implied_ev_ebitda']:.1f}x",    "Implied P/E",           f"{tx['implied_pe']:.1f}x"),
        ("Consideration Mix",   f"{tx['cash_pct']:.0f}% Cash / {tx['stock_pct']:.0f}% Stock",
         "New Shares Issued",   f"{tx['new_shares_issued']:,.1f}M"),
        ("Exchange Ratio",      f"{tx['exchange_ratio']:.4f}x",       "New Debt Raised",       f"${tx['new_debt_issued']:,.0f}M"),
    ]
    for lbl1, val1, lbl2, val2 in summary_rows:
        ws.cell(r, 2, lbl1).font = _font(bold=True); ws.cell(r, 2).border = _BORDER; ws.cell(r, 2).alignment = _align("left")
        ws.cell(r, 3, val1).border = _BORDER; ws.cell(r, 3).alignment = _align("right")
        ws.cell(r, 4, lbl2).font = _font(bold=bool(lbl2)); ws.cell(r, 4).border = _BORDER; ws.cell(r, 4).alignment = _align("left")
        ws.cell(r, 5, val2).border = _BORDER; ws.cell(r, 5).alignment = _align("right")
        r += 1

    r += 1
    # ── EPS impact block ──────────────────────────────────────────────────────
    _hdr(ws, r, 2, 5, "EPS ACCRETION / (DILUTION) — KEY OUTPUTS", bg=_DARK_B, fg=_WHITE)
    r += 1

    ws.row_dimensions[r].height = 16
    for col, hdr in [(2, ""), (3, "Year 1"), (4, "Year 2"), (5, "Year 3")]:
        _hdr(ws, r, col, col, hdr, bg=_HDR_BG, fg="000000")
    r += 1

    def _acc_color(pct):
        if pct >= 0:
            return _GRN, _GRN_BG
        return _RED, _RED_BG

    for label, key in [("GAAP EPS (Acq. Standalone)", "acq_eps_standalone"),
                        ("GAAP Pro Forma EPS",          "pf_eps_gaap"),
                        ("GAAP Accretion / (Dilution)", "gaap_accretion_pct"),
                        ("Cash Pro Forma EPS",          "pf_eps_cash"),
                        ("Cash Accretion / (Dilution)", "cash_accretion_pct")]:
        _label(ws, r, 2, label, bold=(key in ("gaap_accretion_pct", "cash_accretion_pct")))
        for ci, yr in enumerate(pf["years"], 3):
            v = yr[key]
            if key in ("gaap_accretion_pct", "cash_accretion_pct"):
                fg, bg = _acc_color(v)
                _val(ws, r, ci, v / 100, fmt="pct", color=fg, bold=True, bg=bg)
            else:
                _val(ws, r, ci, v, fmt="eps")
        r += 1

    r += 1
    _hdr(ws, r, 2, 5, "SYNERGY & VALUATION CONTEXT", bg=_DARK_B, fg=_WHITE)
    r += 1

    ctx_rows = [
        ("Run-Rate Synergies (After-Tax)", f"${syn['total_aftertax_runrate']:,.0f}M/yr"),
        ("Synergy NPV",                    f"${syn['synergy_npv']:,.0f}M"),
        ("Premium to 52-Week High",        f"{tx['implied_premium_52wk_high']:.1f}%"),
        (f"Target Analyst Price Target",   f"${tgt.analyst_target:.2f}" if tgt.analyst_target else "N/A"),
    ]
    for lbl, val in ctx_rows:
        ws.cell(r, 2, lbl).font = _font(); ws.cell(r, 2).border = _BORDER; ws.cell(r, 2).alignment = _align("left")
        ws.cell(r, 3, val).border = _BORDER; ws.cell(r, 3).alignment = _align("right")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        r += 1


# ── Tab 2: Assumptions ────────────────────────────────────────────────────────

def _build_assumptions(ws, acq: MACompanyData, tgt: MACompanyData, tx: dict, syn: dict):
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 36, 2: 18, 3: 4, 4: 36, 5: 18})

    r = 1
    _hdr(ws, r, 1, 2, "ACQUIRER ASSUMPTIONS", bg=_DARK_B, fg=_WHITE); r += 1
    acq_rows = [
        ("Share Price",            acq.lbo.share_price,        "eps",     "Market"),
        ("Diluted Shares (M)",     acq.diluted_shares,         "number",  "yfinance"),
        ("Market Cap ($M)",        acq.lbo.market_cap,         "dollar0", "Derived"),
        ("LTM Revenue ($M)",       acq.lbo.ltm_revenue,        "dollar0", "FMP/yfinance"),
        ("LTM EBITDA ($M)",        acq.lbo.ltm_ebitda,         "dollar0", "FMP/yfinance"),
        ("LTM Net Income ($M)",    acq.net_income,             "dollar0", "yfinance"),
        ("LTM EPS (Diluted)",      acq.diluted_eps,            "eps",     "yfinance"),
        ("P/E Multiple",           acq.pe_multiple,            "mult",    "yfinance"),
        ("Net Debt ($M)",          acq.lbo.net_debt,           "dollar0", "yfinance"),
        ("Cash ($M)",              acq.lbo.cash,               "dollar0", "yfinance"),
        ("Tax Rate",               acq.lbo.ltm_tax_rate,       "pct",     "FMP/yfinance"),
        ("Credit Rating Proxy",    acq.credit_rating_proxy,    "text",    "Derived"),
        ("Cost of New Debt",       tx["new_debt_rate"],        "pct",     "Derived"),
    ]
    for lbl, val, fmt, src in acq_rows:
        _label(ws, r, 1, lbl)
        if fmt == "text":
            ws.cell(r, 2, val).border = _BORDER
        else:
            _input(ws, r, 2, val, fmt=fmt)
        ws.cell(r, 3, "")
        r += 1

    r = 1
    _hdr(ws, r, 4, 5, "TARGET ASSUMPTIONS", bg=_DARK_B, fg=_WHITE); r += 1
    tgt_rows = [
        ("Share Price (Unaffected)",  tgt.lbo.share_price,     "eps",     "Market"),
        ("Diluted Shares (M)",        tgt.diluted_shares,      "number",  "yfinance"),
        ("Market Cap ($M)",           tgt.lbo.market_cap,      "dollar0", "Market"),
        ("LTM Revenue ($M)",          tgt.lbo.ltm_revenue,     "dollar0", "FMP/yfinance"),
        ("LTM EBITDA ($M)",           tgt.lbo.ltm_ebitda,      "dollar0", "FMP/yfinance"),
        ("LTM Net Income ($M)",       tgt.net_income,          "dollar0", "yfinance"),
        ("LTM EPS (Diluted)",         tgt.diluted_eps,         "eps",     "yfinance"),
        ("Net Debt ($M)",             tgt.lbo.net_debt,        "dollar0", "yfinance"),
        ("Book Equity ($M)",          tgt.book_equity,         "dollar0", "yfinance/FMP"),
        ("52-Week High",              tgt.week52_high,         "eps",     "yfinance"),
        ("52-Week Low",               tgt.week52_low,          "eps",     "yfinance"),
        ("Analyst Price Target",      tgt.analyst_target,      "eps",     "FMP"),
        ("Standalone EV/EBITDA",      tgt.lbo.current_ev_ebitda, "mult",  "Derived"),
    ]
    for lbl, val, fmt, src in tgt_rows:
        _label(ws, r, 4, lbl)
        if fmt == "text":
            ws.cell(r, 5, val).border = _BORDER
        else:
            _input(ws, r, 5, val, fmt=fmt)
        r += 1

    # Deal terms block below
    r = max(len(acq_rows), len(tgt_rows)) + 3
    _hdr(ws, r, 1, 5, "DEAL TERMS (INPUTS)", bg=_NAVY, fg=_WHITE); r += 1
    deal_rows = [
        ("Offer Premium (%)",         tx["offer_premium_pct"],         "number"),
        ("Cash Consideration (%)",    tx["cash_pct"],                  "number"),
        ("Stock Consideration (%)",   tx["stock_pct"],                 "number"),
        ("Run-Rate Synergies ($M)",   syn["total_pretax_runrate"],     "dollar0"),
        ("Synergy Realization Y1",    50.0,                            "number"),
        ("Synergy Realization Y2",    75.0,                            "number"),
        ("Synergy Realization Y3",    100.0,                           "number"),
        ("Intangibles Life (yrs)",    10.0,                            "number"),
        ("Foregone Cash Yield",       4.5,                             "number"),
    ]
    for lbl, val, fmt in deal_rows:
        _label(ws, r, 1, lbl); _input(ws, r, 2, val, fmt=fmt)
        r += 1


# ── Tab 3: Transaction Structure ──────────────────────────────────────────────

def _build_transaction(ws, acq: MACompanyData, tgt: MACompanyData, tx: dict):
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 36, 2: 20, 3: 4, 4: 36, 5: 20})

    r = 1
    _hdr(ws, r, 1, 2, "PURCHASE PRICE ANALYSIS", bg=_DARK_B, fg=_WHITE); r += 1
    pp_rows = [
        ("Target Share Price (Unaffected)",  tgt.lbo.share_price,       "eps"),
        ("Offer Premium",                    tx["offer_premium_pct"] / 100, "pct"),
        ("Offer Price Per Share",            tx["offer_price"],          "eps"),
        ("Target Diluted Shares (M)",        tx["tgt_diluted_shares"],   "number"),
        ("Equity Purchase Price ($M)",       tx["total_equity_value"],   "dollar0"),
        ("(+) Net Debt Assumed ($M)",        tx["net_debt_assumed"],     "dollar0"),
        ("Enterprise Value ($M)",            tx["total_ev"],             "dollar0"),
        ("",                                 "",                         ""),
        ("Implied EV / LTM EBITDA",          tx["implied_ev_ebitda"],    "mult"),
        ("Implied EV / LTM Revenue",         tx["implied_ev_rev"],       "mult"),
        ("Implied P/E (LTM)",                tx["implied_pe"],           "mult"),
        ("Premium to 52-Week High",          tx["implied_premium_52wk_high"] / 100 if tx["implied_premium_52wk_high"] != 0 else 0, "pct"),
    ]
    for lbl, val, fmt in pp_rows:
        _label(ws, r, 1, lbl)
        if val != "":
            _val(ws, r, 2, val, fmt=fmt)
        else:
            ws.cell(r, 2).border = _BORDER
        r += 1

    r = 1
    _hdr(ws, r, 4, 5, "SOURCES & USES", bg=_DARK_B, fg=_WHITE); r += 1
    _hdr(ws, r, 4, 4, "Sources", bg=_HDR_BG, fg="000000")
    _hdr(ws, r, 5, 5, "Amount ($M)", bg=_HDR_BG, fg="000000"); r += 1
    sources = [
        ("Cash from Acquirer Balance",  tx["cash_from_balance"]),
        ("New Debt Issued",             tx["new_debt_issued"]),
        ("New Equity Issued (Stock)",   tx["stock_consideration"]),
        ("Total Sources",               tx["cash_from_balance"] + tx["new_debt_issued"] + tx["stock_consideration"]),
    ]
    for i, (lbl, val) in enumerate(sources):
        bold = (i == len(sources) - 1)
        _label(ws, r, 4, lbl, bold=bold)
        _val(ws, r, 5, val, fmt="dollar0", bold=bold)
        r += 1
    r += 1
    _hdr(ws, r, 4, 4, "Uses", bg=_HDR_BG, fg="000000")
    _hdr(ws, r, 5, 5, "Amount ($M)", bg=_HDR_BG, fg="000000"); r += 1
    total_uses = tx["total_equity_value"] + tx["transaction_fees"]
    uses = [
        ("Equity Purchase Price",       tx["total_equity_value"]),
        ("Transaction Fees",            tx["transaction_fees"]),
        ("Total Uses",                  total_uses),
    ]
    for i, (lbl, val) in enumerate(uses):
        bold = (i == len(uses) - 1)
        _label(ws, r, 4, lbl, bold=bold)
        _val(ws, r, 5, val, fmt="dollar0", bold=bold)
        r += 1

    # PPA
    r_ppa = max(len(pp_rows), len(sources) + len(uses) + 5) + 2
    _hdr(ws, r_ppa, 1, 5, "PURCHASE PRICE ALLOCATION (PPA)", bg=_NAVY, fg=_WHITE); r_ppa += 1
    ppa_rows = [
        ("Equity Purchase Price ($M)",        tx["total_equity_value"],          "dollar0"),
        ("(-) Target Book Equity ($M)",        tgt.book_equity,                   "dollar0"),
        ("Purchase Premium Over Book ($M)",    tx["purchase_premium_over_book"],  "dollar0"),
        ("  Allocated to Intangibles (30%)",   tx["intangibles_acquired"],        "dollar0"),
        ("  Allocated to Goodwill (70%)",      tx["goodwill"],                    "dollar0"),
        ("Annual Intangibles Amortization",    tx["intang_amort_annual"],         "dollar0"),
        ("After-Tax Intangibles Amort.",       tx["intang_amort_aftertax"],       "dollar0"),
    ]
    for lbl, val, fmt in ppa_rows:
        bold = "Goodwill" in lbl or "Amort" in lbl
        _label(ws, r_ppa, 1, lbl, bold=bold)
        _val(ws, r_ppa, 2, val, fmt=fmt, bold=bold)
        ws.cell(r_ppa, 3).border = _BORDER
        ws.cell(r_ppa, 4).border = _BORDER
        ws.cell(ws.max_row, 5).border = _BORDER
        r_ppa += 1


# ── Tab 4 & 5: Acquirer/Target Standalone ─────────────────────────────────────

def _build_standalone(ws, co: MACompanyData, label: str, pf: dict, is_acquirer=True):
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 36, 2: 20, 3: 20, 4: 20})

    r = 1
    _hdr(ws, r, 1, 4, f"{label.upper()} STANDALONE OVERVIEW", bg=_DARK_B, fg=_WHITE); r += 1

    # Key financial metrics
    items = [
        ("Share Price",             co.lbo.share_price,       "eps"),
        ("Market Cap ($M)",         co.lbo.market_cap,        "dollar0"),
        ("EV ($M)",                 co.lbo.current_ev,        "dollar0"),
        ("EV / EBITDA",             co.lbo.current_ev_ebitda, "mult"),
        ("P/E Multiple",            co.pe_multiple,            "mult"),
        ("LTM Revenue ($M)",        co.lbo.ltm_revenue,       "dollar0"),
        ("LTM EBITDA ($M)",         co.lbo.ltm_ebitda,        "dollar0"),
        ("EBITDA Margin",           co.lbo.ltm_ebitda_margin, "pct"),
        ("LTM Net Income ($M)",     co.net_income,            "dollar0"),
        ("Net Margin",              co.net_margin,            "pct"),
        ("LTM EPS (Diluted)",       co.diluted_eps,           "eps"),
        ("Diluted Shares (M)",      co.diluted_shares,        "number"),
        ("Net Debt ($M)",           co.lbo.net_debt,          "dollar0"),
        ("Net Debt / EBITDA",       co.net_debt_ebitda,       "mult"),
        ("Credit Rating Proxy",     co.credit_rating_proxy,   "text"),
        ("52-Week High",            co.week52_high,           "eps"),
        ("52-Week Low",             co.week52_low,            "eps"),
        ("Analyst Price Target",    co.analyst_target,        "eps"),
    ]
    for lbl, val, fmt in items:
        _label(ws, r, 1, lbl)
        if fmt == "text":
            ws.cell(r, 2, val).border = _BORDER; ws.cell(r, 2).alignment = _align("right")
        elif val:
            _val(ws, r, 2, val, fmt=fmt)
        else:
            ws.cell(r, 2).border = _BORDER
        r += 1

    # EPS projection (standalone, no deal)
    r += 1
    _hdr(ws, r, 1, 4, "EPS PROJECTION — STANDALONE (7% annual growth assumed)", bg=_DARK_B, fg=_WHITE); r += 1
    _hdr(ws, r, 1, 1, "Metric", bg=_HDR_BG, fg="000000")
    for yi, yr in enumerate([1, 2, 3]):
        _hdr(ws, r, yi + 2, yi + 2, f"Year {yr}", bg=_HDR_BG, fg="000000")
    r += 1
    growth = 0.07 if is_acquirer else 0.06
    for lbl, yr_fn in [
        ("Net Income ($M)", lambda y: co.net_income * (1 + growth) ** y),
        ("EPS (Diluted)",   lambda y: co.diluted_eps * (1 + growth) ** y),
    ]:
        _label(ws, r, 1, lbl)
        for yi, yr in enumerate([1, 2, 3]):
            fmt = "dollar0" if "Income" in lbl else "eps"
            _val(ws, r, yi + 2, yr_fn(yr), fmt=fmt)
        r += 1

    if not is_acquirer:
        r += 1
        _hdr(ws, r, 1, 4, "VALUATION CONTEXT", bg=_DARK_B, fg=_WHITE); r += 1
        if co.analyst_target:
            upside = (co.analyst_target / co.lbo.share_price - 1) * 100
            _label(ws, r, 1, "Analyst Target Upside/Downside")
            _val(ws, r, 2, upside / 100, fmt="pct"); r += 1
        if co.week52_high:
            pct_52h = (co.lbo.share_price / co.week52_high - 1) * 100
            _label(ws, r, 1, "% Off 52-Week High")
            _val(ws, r, 2, pct_52h / 100, fmt="pct"); r += 1


# ── Tab 6: Pro Forma Combined ─────────────────────────────────────────────────

def _build_pro_forma(ws, acq: MACompanyData, tgt: MACompanyData,
                     tx: dict, syn: dict, pf: dict):
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 38, 2: 20, 3: 20, 4: 20})

    r = 1
    _hdr(ws, r, 1, 4, "PRO FORMA COMBINED INCOME STATEMENT", bg=_NAVY, fg=_WHITE); r += 1
    _hdr(ws, r, 1, 1, "", bg=_HDR_BG, fg="000000")
    for yi, yr in enumerate([1, 2, 3]):
        _hdr(ws, r, yi + 2, yi + 2, f"Pro Forma Year {yr}", bg=_HDR_BG, fg="000000")
    r += 1

    rows_spec = [
        ("Acquirer Standalone Net Income ($M)", "acq_ni_standalone",     "dollar0"),
        ("(+) Target Standalone Net Income ($M)","tgt_ni_standalone",    "dollar0"),
        ("(-) Foregone Interest on Cash (AT)",   "foregone_int",         "dollar0"),
        ("(-) New Debt Interest Expense (AT)",   "new_debt_int",         "dollar0"),
        ("(-) Intangibles Amortization (AT)",    "intang_amort_aftertax","dollar0"),
        ("(+) After-Tax Synergies",              "syn_aftertax",         "dollar0"),
        ("= Pro Forma Net Income GAAP ($M)",     "pf_ni_gaap",           "dollar0"),
        ("(+) Intangibles Amort Add-Back (AT)",  "intang_amort_aftertax","dollar0"),
        ("= Pro Forma Net Income Cash ($M)",     "pf_ni_cash",           "dollar0"),
    ]
    for lbl, key, fmt in rows_spec:
        bold = lbl.startswith("=")
        neg  = lbl.startswith("(-)")
        _label(ws, r, 1, lbl, bold=bold)
        for yi, yr in enumerate(pf["years"]):
            v = yr[key]
            if neg:
                _val(ws, r, yi + 2, -abs(v), fmt=fmt, bold=bold)
            else:
                _val(ws, r, yi + 2, v, fmt=fmt, bold=bold)
        r += 1

    r += 1
    _hdr(ws, r, 1, 4, "SHARE COUNT", bg=_DARK_B, fg=_WHITE); r += 1
    sc_rows = [
        ("Acquirer Standalone Shares (M)", acq.diluted_shares, None),
        ("New Shares Issued (M)",          tx["new_shares_issued"], None),
        ("Pro Forma Shares (M)",           tx["pro_forma_shares"], None),
    ]
    for lbl, val, _ in sc_rows:
        bold = "Pro Forma" in lbl
        _label(ws, r, 1, lbl, bold=bold)
        for col in [2, 3, 4]:
            _val(ws, r, col, val, fmt="number", bold=bold)
        r += 1

    r += 1
    _hdr(ws, r, 1, 4, "SYNERGY RAMP ($M after-tax)", bg=_DARK_B, fg=_WHITE); r += 1
    syn_rows = [
        ("Cost Synergies (AT)",     "cost_syn_aftertax"),
        ("Revenue Synergies (AT)",  "rev_syn_aftertax"),
        ("Total Run-Rate Synergies","total_aftertax_runrate"),
        ("Realization %",           None),
        ("Achieved Synergies (AT)", None),
    ]
    for lbl, key in syn_rows:
        _label(ws, r, 1, lbl, bold=(lbl.startswith("Total") or lbl.startswith("Achieved")))
        for yi, (ramp, yr_syn) in enumerate(zip(
            ["50%", "75%", "100%"],
            [syn["syn_aftertax_y1"], syn["syn_aftertax_y2"], syn["syn_aftertax_y3"]]
        )):
            if key in ("cost_syn_aftertax", "rev_syn_aftertax", "total_aftertax_runrate"):
                _val(ws, r, yi + 2, syn[key], fmt="dollar0")
            elif lbl == "Realization %":
                ws.cell(r, yi + 2, ramp).border = _BORDER
                ws.cell(r, yi + 2).alignment = _align("right")
                ws.cell(r, yi + 2).font = _font()
            else:
                _val(ws, r, yi + 2, yr_syn, fmt="dollar0", bold=True)
        r += 1


# ── Tab 7: Accretion/Dilution ─────────────────────────────────────────────────

def _build_accretion(ws, acq: MACompanyData, tx: dict, syn: dict, pf: dict,
                     be_gaap: float, be_cash: float):
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 38, 2: 20, 3: 20, 4: 20})

    r = 1
    _hdr(ws, r, 1, 4, "EPS ACCRETION / (DILUTION) ANALYSIS", bg=_NAVY, fg=_WHITE); r += 1
    _hdr(ws, r, 1, 1, "", bg=_HDR_BG, fg="000000")
    for yi, yr in enumerate([1, 2, 3]):
        _hdr(ws, r, yi + 2, yi + 2, f"Year {yr}", bg=_HDR_BG, fg="000000")
    r += 1

    # GAAP EPS bridge
    _hdr(ws, r, 1, 4, "GAAP EPS BRIDGE", bg=_DARK_B, fg=_WHITE); r += 1
    gaap_rows = [
        ("Acquirer Standalone EPS",        "acq_eps_standalone", "eps"),
        ("GAAP Pro Forma EPS",             "pf_eps_gaap",        "eps"),
        ("GAAP Accretion / (Dilution) $",  "gaap_accretion_dol", "eps"),
        ("GAAP Accretion / (Dilution) %",  "gaap_accretion_pct", "pct"),
    ]
    for lbl, key, fmt in gaap_rows:
        bold = "Accretion" in lbl
        _label(ws, r, 1, lbl, bold=bold)
        for yi, yr in enumerate(pf["years"]):
            v = yr[key]
            is_pct = key.endswith("_pct")
            if is_pct:
                v = v / 100
            if "Accretion" in lbl:
                fg, bg = (_GRN, _GRN_BG) if v >= 0 else (_RED, _RED_BG)
                _val(ws, r, yi + 2, v, fmt=fmt, color=fg, bold=True, bg=bg)
            else:
                _val(ws, r, yi + 2, v, fmt=fmt, bold=bold)
        r += 1

    r += 1
    _hdr(ws, r, 1, 4, "CASH EPS BRIDGE (GAAP + Intangibles Add-Back)", bg=_DARK_B, fg=_WHITE); r += 1
    cash_rows = [
        ("Acquirer Standalone EPS",        "acq_eps_standalone", "eps"),
        ("Cash Pro Forma EPS",             "pf_eps_cash",        "eps"),
        ("Cash Accretion / (Dilution) $",  "cash_accretion_dol", "eps"),
        ("Cash Accretion / (Dilution) %",  "cash_accretion_pct", "pct"),
    ]
    for lbl, key, fmt in cash_rows:
        bold = "Accretion" in lbl
        _label(ws, r, 1, lbl, bold=bold)
        for yi, yr in enumerate(pf["years"]):
            v = yr[key]
            if key.endswith("_pct"):
                v = v / 100
            if "Accretion" in lbl:
                fg, bg = (_GRN, _GRN_BG) if v >= 0 else (_RED, _RED_BG)
                _val(ws, r, yi + 2, v, fmt=fmt, color=fg, bold=True, bg=bg)
            else:
                _val(ws, r, yi + 2, v, fmt=fmt, bold=bold)
        r += 1

    # Break-even analysis
    r += 2
    _hdr(ws, r, 1, 4, "BREAK-EVEN PREMIUM ANALYSIS", bg=_NAVY, fg=_WHITE); r += 1

    def _be_str(val):
        if math.isnan(val):
            return "No breakeven — dilutive at all premiums"
        return f"{val:.1f}%"

    be_rows = [
        ("GAAP Break-Even Premium",  _be_str(be_gaap)),
        ("Cash EPS Break-Even Premium", _be_str(be_cash)),
        ("Deal Offer Premium",       f"{tx['offer_premium_pct']:.0f}%"),
    ]
    for lbl, val in be_rows:
        _label(ws, r, 1, lbl, bold=True)
        cell = ws.cell(r, 2, val)
        cell.border    = _BORDER
        cell.alignment = _align("right")
        cell.font      = _font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1


# ── Tab 8: Sensitivity ────────────────────────────────────────────────────────

def _build_sensitivity(ws, sens: dict):
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 4, 8: 24, 9: 16, 10: 16, 11: 16, 12: 16, 13: 16})

    def _sens_cell(row, col, val, center_val):
        """Color green if accretive, red if dilutive, medium-blue if center."""
        if val > 0:
            bg = _GRN_BG; fg = _GRN
        else:
            bg = _RED_BG; fg = _RED
        bold = False
        # Highlight center cell
        mid_r = len(sens["premiums"]) // 2
        mid_c_t1 = len(sens["syn_pcts"]) // 2
        is_center = (
            sens["premiums"].index(center_val[0]) == mid_r
            and center_val[1] == mid_c_t1
        )
        if is_center:
            bg = "BDD7EE"; bold = True
        _val(ws, row, col, val / 100, fmt="pct", color=fg, bold=bold, bg=bg)

    # Table 1: Accretion/Dilution vs Premium × Synergy %
    r = 1
    _hdr(ws, r, 1, 6, "TABLE 1 — Year 1 GAAP EPS Impact: Offer Premium × Synergy Realization", bg=_NAVY, fg=_WHITE); r += 1
    _hdr(ws, r, 1, 1, "Premium \\ Synergy %", bg=_HDR_BG, fg="000000")
    for ci, sp in enumerate(sens["syn_pcts"]):
        lbl = "Base Case" if sp == 1.0 else f"{sp*100:.0f}% of Base"
        _hdr(ws, r, ci + 2, ci + 2, lbl, bg=_HDR_BG, fg="000000")
    r += 1
    base_prem = sens["premiums"][len(sens["premiums"]) // 2]
    for ri, prem in enumerate(sens["premiums"]):
        _label(ws, r, 1, f"{prem:.0f}% Premium", bold=(prem == base_prem))
        for ci, sp in enumerate(sens["syn_pcts"]):
            v = sens["table1"][ri][ci]
            _sens_cell(r, ci + 2, v, (prem, ci))
        r += 1

    r += 2
    # Table 2: Accretion/Dilution vs Premium × Cash %
    _hdr(ws, r, 1, 6, "TABLE 2 — Year 1 GAAP EPS Impact: Offer Premium × Cash % Consideration", bg=_NAVY, fg=_WHITE); r += 1
    _hdr(ws, r, 1, 1, "Premium \\ Cash %", bg=_HDR_BG, fg="000000")
    for ci, cp in enumerate(sens["cash_pcts"]):
        _hdr(ws, r, ci + 2, ci + 2, f"{cp:.0f}% Cash", bg=_HDR_BG, fg="000000")
    r += 1
    for ri, prem in enumerate(sens["premiums"]):
        _label(ws, r, 1, f"{prem:.0f}% Premium", bold=(prem == base_prem))
        for ci, cp in enumerate(sens["cash_pcts"]):
            v = sens["table2"][ri][ci]
            fg = _GRN if v >= 0 else _RED
            bg = _GRN_BG if v >= 0 else _RED_BG
            _val(ws, r, ci + 2, v / 100, fmt="pct", color=fg, bg=bg)
        r += 1

    r += 2
    # Table 3: Implied multiples
    _hdr(ws, r, 1, 4, "TABLE 3 — Implied Acquisition Multiples vs Premium", bg=_NAVY, fg=_WHITE); r += 1
    _hdr(ws, r, 1, 1, "Premium", bg=_HDR_BG, fg="000000")
    _hdr(ws, r, 2, 2, "EV/EBITDA", bg=_HDR_BG, fg="000000")
    _hdr(ws, r, 3, 3, "EV/Revenue", bg=_HDR_BG, fg="000000")
    _hdr(ws, r, 4, 4, "P/E", bg=_HDR_BG, fg="000000")
    r += 1
    for row in sens["table3"]:
        _label(ws, r, 1, f"{row['premium_pct']:.0f}%")
        _val(ws, r, 2, row["ev_ebitda"], fmt="mult")
        _val(ws, r, 3, row["ev_revenue"], fmt="mult")
        _val(ws, r, 4, row["pe"], fmt="mult")
        r += 1


# ── Master build function ─────────────────────────────────────────────────────

def build_ma_excel(acq: MACompanyData, tgt: MACompanyData,
                   tx: dict, syn: dict, pf: dict, sens: dict,
                   be_gaap: float, be_cash: float,
                   output_path: str) -> None:
    wb = Workbook()

    tabs = [
        ("Cover",          lambda ws: _build_cover(ws, acq, tgt, tx, pf, syn)),
        ("Assumptions",    lambda ws: _build_assumptions(ws, acq, tgt, tx, syn)),
        ("Transaction",    lambda ws: _build_transaction(ws, acq, tgt, tx)),
        ("Acquirer",       lambda ws: _build_standalone(ws, acq, acq.lbo.company_name, pf, is_acquirer=True)),
        ("Target",         lambda ws: _build_standalone(ws, tgt, tgt.lbo.company_name, pf, is_acquirer=False)),
        ("Pro Forma",      lambda ws: _build_pro_forma(ws, acq, tgt, tx, syn, pf)),
        ("Accretion",      lambda ws: _build_accretion(ws, acq, tx, syn, pf, be_gaap, be_cash)),
        ("Sensitivity",    lambda ws: _build_sensitivity(ws, sens)),
    ]

    # Rename the default sheet and add the rest
    ws0 = wb.active
    ws0.title = tabs[0][0]
    tabs[0][1](ws0)

    for name, fn in tabs[1:]:
        ws = wb.create_sheet(title=name)
        fn(ws)

    wb.save(output_path)
