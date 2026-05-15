"""
Excel educator — adds cell comments to an existing workbook via header text matching.
Never uses hardcoded cell addresses. Matches by cell value text (case-insensitive, substring).
"""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.comments import Comment


_COMMENT_AUTHOR = "University of Cincinnati"
_MAX_COMMENT_LEN = 280  # Excel truncates very long comments


def _normalize(text: str) -> str:
    return text.lower().strip().rstrip(":").rstrip("($m)").strip()


def _find_match(cell_text: str, comments: dict) -> str | None:
    """
    Returns the comment text if cell_text matches any key in comments.
    Checks: exact match, normalized match, or key contained in cell_text.
    """
    cell_n = _normalize(cell_text)
    for key, comment in comments.items():
        key_n = _normalize(key)
        if key_n == cell_n:
            return comment
        if key_n in cell_n or cell_n in key_n:
            return comment
    return None


def add_excel_comments(xlsx_path: str, excel_comments: dict, audience: str = "student") -> int:
    """
    Adds comments to cells in the workbook at xlsx_path whose text matches
    keys in excel_comments. Saves the workbook in place.
    Returns number of comments added.
    """
    if not excel_comments:
        return 0

    wb = load_workbook(xlsx_path)
    comments_added = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                cell_text = cell.value.strip()
                if not cell_text or len(cell_text) > 80:
                    continue  # skip long text / numeric labels

                match = _find_match(cell_text, excel_comments)
                if match:
                    body = str(match)[:_MAX_COMMENT_LEN]
                    cell.comment = Comment(body, _COMMENT_AUTHOR)
                    comments_added += 1

    wb.save(xlsx_path)
    return comments_added
