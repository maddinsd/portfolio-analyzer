"""Excel output for the LBO model — 9-tab Goldman-style workbook."""
from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter

# ── Design constants ──────────────────────────────────────────────────────────
_NAVY   = "003366"
_DARK_B = "1F4E79"
_MED_B  = "BDD7EE"
_LT_B   = "D9E1F2"
_WHITE  = "FFFFFF"
_INPUT  = "EBF3FB"   # blue-tinted inputs
_GREEN  = "00B050"
_YELLOW = "FFEB9C"
_RED_F  = "FF0000"
_GRN_BG = "C6EFCE"
_YLW_BG = "FFEB9C"
_RED_BG = "FFC7CE"

# Font colors per convention (formula color coding)
_FC_INPUT   = "0000FF"   # blue  — hardcoded inputs
_FC_FORMULA = "000000"   # black — calculations
_FC_LINK    = "800080"   # purple — same-tab link
_FC_XLINK   = "008000"   # green  — cross-tab link

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOT    = Border(bottom=Side(style="medium", color="003366"))


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=_FC_FORMULA, size=11, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _hdr_style(ws, row: int, col_start: int, col_end: int, label: str,
               fill_color: str = _DARK_B, font_size: int = 11) -> None:
    """Write a section header spanning columns."""
    ws.cell(row, col_start, label).font  = _font(bold=True, color=_WHITE, size=font_size)
    ws.cell(row, col_start).fill        = _fill(fill_color)
    ws.cell(row, col_start).alignment   = Alignment(horizontal="left", vertical="center")
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row, c).fill = _fill(fill_color)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)


def _col_hdr(ws, row: int, cols: list, labels: list) -> None:
    for c, lbl in zip(cols, labels):
        cell = ws.cell(row, c, lbl)
        cell.fill      = _fill(_LT_B)
        cell.font      = _font(bold=True, color=_FC_FORMULA)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _BORDER


def _write(ws, row: int, col: int, value, fmt: str | None = None,
           bold: bool = False, fc: str = _FC_FORMULA,
           fill: str | None = None, align: str = "right") -> None:
    cell = ws.cell(row, col, value)
    cell.font      = _font(bold=bold, color=fc)
    cell.alignment = Alignment(horizontal=align)
    cell.border    = _BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = _fill(fill)


def _input(ws, row: int, col: int, value, fmt: str | None = None) -> None:
    _write(ws, row, col, value, fmt=fmt, fc=_FC_INPUT, fill=_INPUT)


def _label(ws, row: int, col: int, text: str, indent: int = 0,
           bold: bool = False) -> None:
    cell = ws.cell(row, col, " " * indent + text)
    cell.font      = _font(bold=bold, color=_FC_FORMULA)
    cell.alignment = Alignment(horizontal="left")
    cell.border    = _BORDER


def _pct(v: float) -> str:
    """Format float as percentage string."""
    return f"{v*100:.1f}%"


def _m(v: float) -> float:
    """Round to 1 decimal for $M display."""
    return round(v, 1)


def _irr_fmt(v: float) -> str:
    if math.isnan(v) or math.isinf(v) or v < -0.99:
        return "N/M"
    return f"{v*100:.1f}%"


FMT_M    = '#,##0.0"M"'
FMT_B    = '#,##0.0"B"'
FMT_PCT  = "0.0%"
FMT_MULT = '0.0"x"'
FMT_MULT2= '0.00"x"'


# ── Tab 1: Cover ──────────────────────────────────────────────────────────────

def _build_cover(ws, data: dict) -> None:
    inp = data["inputs"]
    tx  = data["transaction"]
    ret = data["returns"]
    a   = data["assumptions"]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # Title banner
    _hdr_style(ws, 1, 1, 4, f"LEVERAGED BUYOUT ANALYSIS — {inp.ticker}", _NAVY, 14)
    ws.row_dimensions[1].height = 30
    _hdr_style(ws, 2, 1, 4, inp.company_name, _DARK_B, 12)

    # Key return metrics (large font)
    ws.row_dimensions[5].height = 40
    for col, (lbl, val, fmt) in enumerate([
        ("Gross IRR",  _irr_fmt(ret["gross_irr"]),  None),
        ("Gross MOIC", f"{ret['gross_moic']:.2f}x",  None),
        ("Net IRR",    _irr_fmt(ret["net_irr"]),    None),
        ("Net MOIC",   f"{ret['net_moic']:.2f}x",    None),
    ], 1):
        hcell = ws.cell(4, col, lbl)
        hcell.font = _font(bold=True, color=_WHITE, size=10)
        hcell.fill = _fill(_DARK_B)
        hcell.alignment = Alignment(horizontal="center", vertical="center")
        hcell.border = _BORDER

        vcell = ws.cell(5, col, val)
        irr_v = ret["gross_irr"] if col in (1, 3) else None
        moic_v= ret["gross_moic"] if col in (2, 4) else None
        color = _GRN_BG
        if irr_v is not None and not math.isnan(irr_v):
            color = _GRN_BG if irr_v >= 0.20 else (_YLW_BG if irr_v >= 0.15 else _RED_BG)
        elif moic_v is not None:
            color = _GRN_BG if moic_v >= 2.5 else (_YLW_BG if moic_v >= 2.0 else _RED_BG)
        vcell.font = _font(bold=True, color=_FC_FORMULA, size=22)
        vcell.fill = _fill(color)
        vcell.alignment = Alignment(horizontal="center", vertical="center")
        vcell.border = _BORDER
    ws.row_dimensions[5].height = 50

    r = 7
    # Transaction summary
    _hdr_style(ws, r, 1, 4, "TRANSACTION SUMMARY", _DARK_B)
    r += 1
    for lbl, val, fmt in [
        ("Entry EV/EBITDA",     f"{a.entry_ev_ebitda:.1f}x",             None),
        ("Exit EV/EBITDA",      f"{a.exit_ev_ebitda:.1f}x",              None),
        ("Entry EV ($M)",       f"${tx['entry_ev']:,.0f}M",              None),
        ("Total Debt ($M)",     f"${tx['total_debt']:,.0f}M",            None),
        ("Sponsor Equity ($M)", f"${tx['sponsor_equity']:,.0f}M",         None),
        ("Debt / EBITDA",       f"{tx['debt_ebitda']:.1f}x",             None),
        ("Blended Interest Rate", f"{tx['blended_rate']*100:.1f}%",      None),
        ("Interest Coverage",   f"{tx['coverage_yr1']:.2f}x",            None),
        ("Hold Period",         f"{a.hold_years} years",                  None),
        ("Exit EBITDA ($M)",    f"${ret['exit_ebitda']:,.0f}M",           None),
        ("Exit EV ($M)",        f"${ret['exit_ev']:,.0f}M",              None),
        ("Debt Paydown ($M)",   f"${ret['debt_paydown']:,.0f}M",          None),
    ]:
        _label(ws, r, 1, lbl, bold=False)
        _write(ws, r, 2, val, align="left")
        r += 1

    # Warnings
    r += 1
    warnings = data["inputs"].warnings if hasattr(data.get("inputs", None), "warnings") else []
    if warnings:
        _hdr_style(ws, r, 1, 4, "⚠ MODEL FLAGS", "FF0000")
        r += 1
        for w in warnings:
            ws.cell(r, 1, w).font = _font(bold=True, color="FF0000")
            ws.cell(r, 1).alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.row_dimensions[r].height = 30
            r += 1

    ws.sheet_view.showGridLines = False


# ── Tab 2: Assumptions ────────────────────────────────────────────────────────

def _build_assumptions(ws, data: dict) -> None:
    inp = data["inputs"]
    a   = data["assumptions"]
    tx  = data["transaction"]

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18

    r = 1
    _hdr_style(ws, r, 1, 2, "LBO MODEL ASSUMPTIONS", _NAVY); r += 1

    sections = [
        ("ENTRY ASSUMPTIONS", [
            ("Entry EV/EBITDA Multiple",     a.entry_ev_ebitda,          FMT_MULT,  True),
            ("Current Trading EV/EBITDA",    inp.current_ev_ebitda,      FMT_MULT,  False),
            ("Transaction Fees",             a.transaction_fees_pct,     FMT_PCT,   True),
            ("Management Rollover (% equity)",a.mgmt_rollover_pct,       FMT_PCT,   True),
        ]),
        ("CAPITAL STRUCTURE (% of Entry EV)", [
            ("Senior Secured Term Loan B",   a.tlb_pct,                  FMT_PCT,   True),
            ("Senior Unsecured Notes",       a.senior_notes_pct,         FMT_PCT,   True),
            ("Subordinated Debt",            a.sub_debt_pct,             FMT_PCT,   True),
            ("Total Debt",                   a.tlb_pct+a.senior_notes_pct+a.sub_debt_pct, FMT_PCT, False),
            ("Sponsor + Mgmt Equity",        1-(a.tlb_pct+a.senior_notes_pct+a.sub_debt_pct), FMT_PCT, False),
        ]),
        ("INTEREST RATES", [
            ("TLB Rate (SOFR + 300bps, cap 8.5%)",  a.tlb_rate,           FMT_PCT,   True),
            ("Senior Unsecured Notes Rate",           a.senior_notes_rate,  FMT_PCT,   True),
            ("Subordinated Debt Rate",                a.sub_debt_rate,      FMT_PCT,   True),
            ("Revolver Commitment Fee",               a.revolver_commit_fee,FMT_PCT,   True),
            ("TLB Mandatory Amortization",            a.tlb_amort_pct,      FMT_PCT,   True),
        ]),
        ("LTM OPERATING METRICS (INPUT)", [
            ("Revenue ($M)",                 inp.ltm_revenue,             FMT_M,     True),
            ("EBITDA ($M)",                  inp.ltm_ebitda,              FMT_M,     True),
            ("EBITDA Margin",                inp.ltm_ebitda_margin,       FMT_PCT,   False),
            ("D&A ($M)",                     inp.ltm_da,                  FMT_M,     True),
            ("CapEx ($M)",                   inp.ltm_capex,               FMT_M,     True),
            ("CapEx % Revenue",              inp.capex_pct_rev,           FMT_PCT,   False),
            ("Net Working Capital % Rev",    inp.nwc_pct_rev,             FMT_PCT,   True),
            ("Effective Tax Rate",           inp.ltm_tax_rate,            FMT_PCT,   True),
            ("FCF Conversion (FCF/EBITDA)",  inp.fcf_conversion,          FMT_PCT,   False),
        ]),
        ("REVENUE GROWTH ASSUMPTIONS", [
            ("Year 1 (Analyst Consensus)",   a.rev_growth[0],             FMT_PCT,   True),
            ("Year 2 (Analyst Consensus)",   a.rev_growth[1],             FMT_PCT,   True),
            ("Year 3 (Sector Median)",       a.rev_growth[2],             FMT_PCT,   True),
            ("Year 4 (Sector Median)",       a.rev_growth[3],             FMT_PCT,   True),
            ("Year 5 (Sector Median)",       a.rev_growth[4],             FMT_PCT,   True),
        ]),
        ("EBITDA MARGIN", [
            ("LTM Margin (Base)",            a.ebitda_margin_y0,          FMT_PCT,   False),
            ("Annual Expansion (Ops Improvement)",a.ebitda_margin_exp,    FMT_PCT,   True),
        ]),
        ("EXIT ASSUMPTIONS", [
            ("Exit Year",                    a.exit_year,                 "0",       True),
            ("Exit EV/EBITDA Multiple",      a.exit_ev_ebitda,            FMT_MULT,  True),
            ("Mgmt Promote MOIC Threshold",  a.mgmt_promote_moic,         FMT_MULT2, True),
            ("Mgmt Promote % of Upside",     a.mgmt_promote_pct,          FMT_PCT,   True),
        ]),
        ("NET RETURN ASSUMPTIONS (LP-LEVEL)", [
            ("Annual Mgmt Fee (% committed)",a.mgmt_fee_pct,              FMT_PCT,   True),
            ("Carried Interest %",           a.carry_pct,                 FMT_PCT,   True),
        ]),
    ]

    for section_name, rows in sections:
        _hdr_style(ws, r, 1, 2, section_name); r += 1
        for lbl, val, fmt, is_input in rows:
            _label(ws, r, 1, lbl, indent=2)
            if is_input:
                _input(ws, r, 2, val, fmt)
            else:
                _write(ws, r, 2, val, fmt=fmt, fc=_FC_LINK)
            r += 1

    ws.sheet_view.showGridLines = False


# ── Tab 3: Transaction Structure ──────────────────────────────────────────────

def _build_transaction(ws, data: dict) -> None:
    inp = data["inputs"]
    tx  = data["transaction"]
    a   = data["assumptions"]

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    r = 1
    _hdr_style(ws, r, 1, 4, "SOURCES & USES OF FUNDS", _NAVY, 13); r += 1

    # Column headers
    for c, lbl in [(1, ""), (2, "Amount ($M)"), (3, "% of Total"), (4, "x EBITDA")]:
        _write(ws, r, c, lbl, bold=True, fill=_LT_B,
               align="center" if c > 1 else "left")
    r += 1

    # SOURCES
    _hdr_style(ws, r, 1, 4, "SOURCES OF FUNDS", _DARK_B); r += 1
    ev = tx["entry_ev"]
    ebitda = inp.ltm_ebitda

    sources = [
        ("Senior Secured Term Loan B",   tx["tlb"],          tx["tlb"]/tx["total_uses"],           tx["tlb"]/ebitda if ebitda else 0),
        ("Senior Unsecured Notes",        tx["senior_notes"], tx["senior_notes"]/tx["total_uses"],   tx["senior_notes"]/ebitda if ebitda else 0),
        ("Subordinated Debt",             tx["sub_debt"],     tx["sub_debt"]/tx["total_uses"],       tx["sub_debt"]/ebitda if ebitda else 0),
        ("Management Rollover Equity",    tx["mgmt_rollover"],tx["mgmt_rollover"]/tx["total_uses"],  tx["mgmt_rollover"]/ebitda if ebitda else 0),
        ("Sponsor Equity",                tx["sponsor_equity"],tx["sponsor_equity"]/tx["total_uses"],tx["sponsor_equity"]/ebitda if ebitda else 0),
    ]
    total_sources = tx["total_uses"]

    for lbl, amt, pct, lev in sources:
        _label(ws, r, 1, lbl, indent=2)
        _write(ws, r, 2, amt, fmt=FMT_M, fc=_FC_XLINK)
        _write(ws, r, 3, pct, fmt=FMT_PCT)
        _write(ws, r, 4, lev, fmt=FMT_MULT)
        r += 1

    _label(ws, r, 1, "Total Sources", bold=True)
    _write(ws, r, 2, total_sources, fmt=FMT_M, bold=True, fill=_MED_B)
    _write(ws, r, 3, 1.0, fmt=FMT_PCT, bold=True, fill=_MED_B)
    _write(ws, r, 4, total_sources/ebitda if ebitda else 0, fmt=FMT_MULT, bold=True, fill=_MED_B)
    r += 2

    # USES
    _hdr_style(ws, r, 1, 4, "USES OF FUNDS", _DARK_B); r += 1

    uses = [
        ("Purchase Enterprise Value",     ev,                     ev/total_sources,               ev/ebitda if ebitda else 0),
        ("Transaction & Financing Fees",  tx["transaction_fees"], tx["transaction_fees"]/total_sources, tx["transaction_fees"]/ebitda if ebitda else 0),
    ]
    for lbl, amt, pct, lev in uses:
        _label(ws, r, 1, lbl, indent=2)
        _write(ws, r, 2, amt, fmt=FMT_M, fc=_FC_XLINK)
        _write(ws, r, 3, pct, fmt=FMT_PCT)
        _write(ws, r, 4, lev, fmt=FMT_MULT)
        r += 1

    _label(ws, r, 1, "Total Uses", bold=True)
    _write(ws, r, 2, total_sources, fmt=FMT_M, bold=True, fill=_MED_B)
    _write(ws, r, 3, 1.0, fmt=FMT_PCT, bold=True, fill=_MED_B)
    r += 2

    # Capital structure metrics
    _hdr_style(ws, r, 1, 4, "CAPITAL STRUCTURE METRICS", _DARK_B); r += 1
    metrics = [
        ("Entry EV ($M)",                   ev,                              FMT_M),
        ("Entry EV / LTM EBITDA",           a.entry_ev_ebitda,               FMT_MULT),
        ("Total Debt ($M)",                  tx["total_debt"],                FMT_M),
        ("Total Debt / LTM EBITDA",         tx["debt_ebitda"],               FMT_MULT),
        ("Blended Interest Rate",            tx["blended_rate"],              FMT_PCT),
        ("Estimated Year 1 Interest ($M)",   tx["interest_yr1"],              FMT_M),
        ("Interest Coverage (EBITDA / Int)", tx["coverage_yr1"],              FMT_MULT),
        ("Debt % of Total Capitalization",   tx["debt_pct"],                  FMT_PCT),
        ("Equity % of Total Capitalization", tx["equity_pct"],                FMT_PCT),
    ]
    for lbl, val, fmt in metrics:
        _label(ws, r, 1, lbl, indent=2)
        _write(ws, r, 2, val, fmt=fmt)
        r += 1

    # Coverage flag
    cov = tx["coverage_yr1"]
    flag_color = _GRN_BG if cov >= 2.0 else (_YLW_BG if cov >= 1.5 else _RED_BG)
    ws.cell(r - len(metrics) + 6, 2).fill = _fill(flag_color)

    ws.sheet_view.showGridLines = False


# ── Tab 4: Income Statement ───────────────────────────────────────────────────

def _build_income_statement(ws, data: dict) -> None:
    inp  = data["inputs"]
    stmts= data["statements"]
    is_r = stmts["income_stmt"]
    a    = data["assumptions"]

    ws.column_dimensions["A"].width = 34
    col_w = [14] * (a.hold_years + 1)
    for i, w in enumerate(col_w):
        ws.column_dimensions[get_column_letter(i + 2)].width = w

    r = 1
    _hdr_style(ws, r, 1, a.hold_years + 2, "PROJECTED INCOME STATEMENT ($M)", _NAVY, 13); r += 1

    # Column headers: LTM + Years
    hdrs = ["LTM Actual"] + [f"Year {y}" for y in range(1, a.hold_years + 1)]
    _col_hdr(ws, r, list(range(2, a.hold_years + 3)), hdrs)
    r += 1

    # LTM column values
    ltm = {
        "revenue": inp.ltm_revenue, "ebitda": inp.ltm_ebitda,
        "ebitda_margin": inp.ltm_ebitda_margin, "da": inp.ltm_da,
        "ebit": inp.ltm_ebit, "interest_exp": inp.ltm_interest_exp,
        "ebt": inp.ltm_ebit - inp.ltm_interest_exp,
        "tax": (inp.ltm_ebit - inp.ltm_interest_exp) * inp.ltm_tax_rate,
        "net_income": (inp.ltm_ebit - inp.ltm_interest_exp) * (1 - inp.ltm_tax_rate),
    }

    def _is_row(label: str, key: str, fmt: str,
                is_margin: bool = False, subtotal: bool = False, total: bool = False) -> None:
        nonlocal r
        _label(ws, r, 1, label, indent=(0 if total or subtotal else 2), bold=(total or subtotal))
        ltm_val = ltm.get(key, 0)
        _write(ws, r, 2, ltm_val, fmt=fmt, fc=_FC_INPUT, fill=_INPUT if not is_margin else None)
        for c, yr in enumerate(is_r, 3):
            val = yr.get(key, 0)
            fill_c = _MED_B if total else None
            _write(ws, r, c, val, fmt=fmt, bold=total,
                   fc=_FC_XLINK, fill=fill_c)
        r += 1

    _hdr_style(ws, r, 1, a.hold_years + 2, "REVENUE & EBITDA", _DARK_B); r += 1
    _is_row("Revenue",          "revenue",        FMT_M)
    _is_row("  % Growth",       "_rev_growth",    FMT_PCT)  # placeholder, write manually
    # Write revenue growth row manually
    row_growth = r - 1
    for c, yr_idx in enumerate(range(len(is_r)), 3):
        if yr_idx == 0:
            g = a.rev_growth[0]
        else:
            g = a.rev_growth[yr_idx] if yr_idx < len(a.rev_growth) else a.rev_growth[-1]
        _write(ws, row_growth, c, g, fmt=FMT_PCT, fc=_FC_XLINK)
    _write(ws, row_growth, 2, None, fmt=FMT_PCT, fc=_FC_INPUT)

    _is_row("EBITDA",           "ebitda",         FMT_M, subtotal=True)
    _is_row("  EBITDA Margin",  "ebitda_margin",  FMT_PCT, is_margin=True)
    _is_row("  D&A",            "da",             FMT_M)
    _is_row("EBIT",             "ebit",           FMT_M, subtotal=True)

    _hdr_style(ws, r, 1, a.hold_years + 2, "BELOW THE LINE", _DARK_B); r += 1
    _is_row("  Interest Expense", "interest_exp", FMT_M)
    _is_row("EBT",               "ebt",           FMT_M, subtotal=True)
    _is_row("  Income Taxes",    "tax",            FMT_M)
    _is_row("Net Income",        "net_income",     FMT_M, total=True)

    # EBITDA margin row — color green/yellow/red by coverage
    ws.sheet_view.showGridLines = False


# ── Tab 5: Balance Sheet ──────────────────────────────────────────────────────

def _build_balance_sheet(ws, data: dict) -> None:
    inp   = data["inputs"]
    stmts = data["statements"]
    bs_r  = stmts["balance_sheet"]
    tx    = data["transaction"]
    a     = data["assumptions"]

    ws.column_dimensions["A"].width = 34
    for i in range(a.hold_years + 2):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    r = 1
    _hdr_style(ws, r, 1, a.hold_years + 2, "PROJECTED BALANCE SHEET ($M)", _NAVY, 13); r += 1
    hdrs = ["At Close"] + [f"Year {y}" for y in range(1, a.hold_years + 1)]
    _col_hdr(ws, r, list(range(2, a.hold_years + 3)), hdrs); r += 1

    # At Close column (post-LBO)
    close = {
        "cash":          max(inp.cash * 0.5, inp.ltm_revenue * 0.01),
        "ar_inventory":  inp.accounts_receivable + inp.inventory,
        "ppe_net":       inp.ppe_net,
        "goodwill":      max(0, tx["entry_ev"] - inp.book_equity),
        "total_assets":  0,
        "accounts_payable": inp.accounts_payable,
        "debt":          tx["total_debt"],
        "total_liab":    0,
        "total_equity":  tx["equity_total"],
        "bs_check":      0,
    }
    close["total_assets"] = close["cash"] + close["ar_inventory"] + close["ppe_net"] + close["goodwill"]
    close["total_liab"]   = close["accounts_payable"] + close["debt"]

    def _bs_row(label: str, key: str, fmt: str = FMT_M,
                subtotal: bool = False, total: bool = False,
                indent: int = 2, check: bool = False) -> None:
        nonlocal r
        _label(ws, r, 1, label, indent=indent, bold=(total or subtotal))
        cl_val = close.get(key, 0)
        _write(ws, r, 2, cl_val, fmt=fmt, fc=_FC_INPUT, fill=_INPUT if not check else None)
        for c, yr in enumerate(bs_r, 3):
            val = yr.get(key, 0)
            fill_c = _MED_B if total else (_RED_BG if (check and abs(val) > 0.5) else (_GRN_BG if check else None))
            _write(ws, r, c, val, fmt=fmt, bold=(total or check), fc=_FC_XLINK, fill=fill_c)
        r += 1

    _hdr_style(ws, r, 1, a.hold_years + 2, "ASSETS", _DARK_B); r += 1
    _bs_row("  Cash & Equivalents",       "cash",          FMT_M)
    _bs_row("  Accounts Receivable & Inventory", "ar_inventory", FMT_M)
    _bs_row("  PP&E, Net",                "ppe_net",       FMT_M)
    _bs_row("  Goodwill (Acquisition Premium)", "goodwill", FMT_M)
    _bs_row("Total Assets",               "total_assets",  FMT_M, total=True)

    _hdr_style(ws, r, 1, a.hold_years + 2, "LIABILITIES & EQUITY", _DARK_B); r += 1
    _bs_row("  Accounts Payable",         "accounts_payable", FMT_M)
    _bs_row("  Total Debt",               "debt",          FMT_M)
    _bs_row("Total Liabilities",          "total_liab",    FMT_M, subtotal=True)
    _bs_row("  Total Equity",             "total_equity",  FMT_M)

    r += 1
    _hdr_style(ws, r, 1, a.hold_years + 2, "BALANCE CHECK (must = 0)", "FF0000"); r += 1
    _bs_row("Assets − Liabilities − Equity", "bs_check", FMT_M, check=True, indent=2)

    ws.sheet_view.showGridLines = False


# ── Tab 6: Cash Flow Statement ────────────────────────────────────────────────

def _build_cash_flow(ws, data: dict) -> None:
    stmts = data["statements"]
    cf_r  = stmts["cash_flow"]
    bs_r  = stmts["balance_sheet"]
    a     = data["assumptions"]

    ws.column_dimensions["A"].width = 36
    for i in range(a.hold_years + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    r = 1
    _hdr_style(ws, r, 1, a.hold_years + 1, "FREE CASH FLOW STATEMENT ($M)", _NAVY, 13); r += 1
    hdrs = [f"Year {y}" for y in range(1, a.hold_years + 1)]
    _col_hdr(ws, r, list(range(2, a.hold_years + 2)), hdrs); r += 1

    def _cf_row(label, key, fmt=FMT_M, subtotal=False, total=False, neg=False, indent=2):
        nonlocal r
        _label(ws, r, 1, label, indent=indent, bold=(total or subtotal))
        for c, yr in enumerate(cf_r, 2):
            val = yr.get(key, 0) * (-1 if neg else 1)
            fill_c = _MED_B if total else None
            _write(ws, r, c, val, fmt=fmt, bold=total, fc=_FC_XLINK, fill=fill_c)
        r += 1

    _hdr_style(ws, r, 1, a.hold_years + 1, "OPERATING CASH FLOW", _DARK_B); r += 1
    _cf_row("  Net Income",                   "net_income")
    _cf_row("  (+) Depreciation & Amortization","da")
    _cf_row("  (−) Change in Net Working Capital","nwc_change", neg=True)
    _cf_row("Cash from Operations",           "cfo",         subtotal=True, indent=0)

    _hdr_style(ws, r, 1, a.hold_years + 1, "INVESTING ACTIVITIES", _DARK_B); r += 1
    _cf_row("  (−) Capital Expenditures",     "capex",       neg=True)

    _hdr_style(ws, r, 1, a.hold_years + 1, "FREE CASH FLOW", _DARK_B); r += 1
    _cf_row("Levered Free Cash Flow",         "levered_fcf", total=True, indent=0)

    _hdr_style(ws, r, 1, a.hold_years + 1, "FINANCING ACTIVITIES (DEBT)", _DARK_B); r += 1
    _cf_row("  (−) Mandatory Amortization + Sweep", "total_paydown", neg=True)

    _hdr_style(ws, r, 1, a.hold_years + 1, "NET CASH FLOW", _DARK_B); r += 1
    _cf_row("Net Cash to / (from) Balance Sheet", "net_cf", total=True, indent=0)

    r += 1
    _hdr_style(ws, r, 1, a.hold_years + 1, "CASH BALANCE ROLL", _DARK_B); r += 1
    _label(ws, r, 1, "  Beginning Cash", bold=False)
    prev_cash = data["inputs"].cash
    for c, yr in enumerate(bs_r, 2):
        beg = prev_cash
        _write(ws, r, c, beg, fmt=FMT_M, fc=_FC_XLINK)
        prev_cash = yr["cash"]
    r += 1
    _label(ws, r, 1, "  Ending Cash", bold=False)
    for c, yr in enumerate(bs_r, 2):
        _write(ws, r, c, yr["cash"], fmt=FMT_M, fc=_FC_XLINK)
    r += 1

    ws.sheet_view.showGridLines = False


# ── Tab 7: Debt Schedule ──────────────────────────────────────────────────────

def _build_debt_schedule(ws, data: dict) -> None:
    stmts = data["statements"]
    ds    = stmts["debt_sched2"]["schedule"]
    tx    = data["transaction"]
    a     = data["assumptions"]

    ws.column_dimensions["A"].width = 38
    for i in range(a.hold_years + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    r = 1
    _hdr_style(ws, r, 1, a.hold_years + 1, "DEBT SCHEDULE ($M)", _NAVY, 13); r += 1
    hdrs = ["At Close"] + [f"Year {y}" for y in range(1, a.hold_years + 1)]
    _col_hdr(ws, r, list(range(2, a.hold_years + 3)), hdrs); r += 1

    def _ds_section(title: str, tranche_key: str, orig_balance: float,
                    rate: float, amort_label: str = "Mandatory Amortization") -> None:
        nonlocal r
        _hdr_style(ws, r, 1, a.hold_years + 2, title, _DARK_B); r += 1

        rows = [
            ("  Beginning Balance",   "beg",   True),
            (f"  ({amort_label})",    "amort", False),
            ("  (Cash Sweep)",        "sweep", False),
            ("  Interest Expense",    "int",   False),
            ("  Ending Balance",      "end",   True),
        ]
        for lbl, key_suffix, bold in rows:
            _label(ws, r, 1, lbl, bold=bold)
            _write(ws, r, 2, orig_balance if key_suffix == "beg" else 0,
                   fmt=FMT_M, fc=_FC_INPUT, fill=_INPUT)
            for c, yr in enumerate(ds, 3):
                val = yr.get(f"{tranche_key}_{key_suffix}", 0)
                fill_c = _MED_B if bold else None
                _write(ws, r, c, val, fmt=FMT_M, fc=_FC_XLINK, fill=fill_c)
            r += 1

        _label(ws, r, 1, f"  Rate: {rate*100:.2f}%", indent=0)
        r += 1

    _ds_section("TERM LOAN B",
                "tlb", tx["tlb"], a.tlb_rate,
                "1% Mandatory Amort + Sweep")
    _ds_section("SENIOR UNSECURED NOTES (Bullet)",
                "sn", tx["senior_notes"], a.senior_notes_rate,
                "None (Bullet at Maturity)")
    _ds_section("SUBORDINATED DEBT (Bullet)",
                "sub", tx["sub_debt"], a.sub_debt_rate,
                "None (Bullet at Maturity)")

    r += 1
    _hdr_style(ws, r, 1, a.hold_years + 2, "TOTAL DEBT SUMMARY", _NAVY); r += 1
    total_rows = [
        ("Total Beginning Debt",  "total_debt_beg", True),
        ("Total Debt Paydown",    "total_paydown",  False),
        ("Total Interest Expense","total_int",      False),
        ("Total Ending Debt",     "total_debt_end", True),
    ]
    for lbl, key, bold in total_rows:
        _label(ws, r, 1, lbl, bold=bold)
        _write(ws, r, 2, tx["total_debt"] if key == "total_debt_beg" else 0,
               fmt=FMT_M, fc=_FC_INPUT, fill=_INPUT)
        for c, yr in enumerate(ds, 3):
            _write(ws, r, c, yr[key], fmt=FMT_M, bold=bold,
                   fc=_FC_XLINK, fill=_MED_B if bold else None)
        r += 1

    ws.sheet_view.showGridLines = False


# ── Tab 8: Returns ────────────────────────────────────────────────────────────

def _build_returns(ws, data: dict) -> None:
    ret = data["returns"]
    tx  = data["transaction"]
    a   = data["assumptions"]
    inp = data["inputs"]

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22

    r = 1
    _hdr_style(ws, r, 1, 2, "RETURNS ANALYSIS", _NAVY, 13); r += 1

    # Exit bridge
    _hdr_style(ws, r, 1, 2, f"EXIT BRIDGE (YEAR {a.exit_year})", _DARK_B); r += 1
    bridge = [
        ("Exit EBITDA ($M)",          ret["exit_ebitda"],       FMT_M),
        ("Exit EV/EBITDA Multiple",   ret["exit_ev_ebitda"],    FMT_MULT),
        ("Exit Enterprise Value ($M)",ret["exit_ev"],           FMT_M),
        ("Less: Remaining Debt ($M)", -ret["exit_debt"],        FMT_M),
        ("Exit Equity Value ($M)",    ret["exit_equity_total"], FMT_M),
    ]
    for lbl, val, fmt in bridge:
        _label(ws, r, 1, lbl, indent=2)
        _write(ws, r, 2, val, fmt=fmt, fc=_FC_FORMULA)
        r += 1

    r += 1
    _hdr_style(ws, r, 1, 2, "EQUITY WATERFALL", _DARK_B); r += 1
    waterfall = [
        ("Exit Equity Value ($M)",          ret["exit_equity_total"],  FMT_M),
        ("Mgmt Promote ($M)",               ret["promote"],            FMT_M),
        ("Sponsor Proceeds ($M)",           ret["sponsor_proceeds"],   FMT_M),
        ("Sponsor Equity Invested ($M)",    tx["sponsor_equity"],      FMT_M),
        ("Mgmt Rollover Invested ($M)",     tx["mgmt_rollover"],       FMT_M),
        ("Mgmt Proceeds ($M)",              ret["mgmt_proceeds"],      FMT_M),
    ]
    for lbl, val, fmt in waterfall:
        _label(ws, r, 1, lbl, indent=2)
        _write(ws, r, 2, val, fmt=fmt, fc=_FC_FORMULA)
        r += 1

    r += 1
    _hdr_style(ws, r, 1, 2, "KEY RETURN METRICS", _NAVY); r += 1
    irr = ret["gross_irr"]
    moic= ret["gross_moic"]

    gross_irr_color = _GRN_BG if (not math.isnan(irr) and irr >= 0.20) else (_YLW_BG if (not math.isnan(irr) and irr >= 0.15) else _RED_BG)
    gross_moic_color= _GRN_BG if moic >= 2.5 else (_YLW_BG if moic >= 2.0 else _RED_BG)

    returns_rows = [
        ("Gross IRR",                _irr_fmt(ret["gross_irr"]),     None, gross_irr_color),
        ("Gross MOIC",               f"{ret['gross_moic']:.2f}x",    None, gross_moic_color),
        ("Net IRR (after fees+carry)",_irr_fmt(ret["net_irr"]),      None, None),
        ("Net MOIC",                 f"{ret['net_moic']:.2f}x",      None, None),
        ("Cash-on-Cash Return",      f"{ret['coc']:.2f}x",           None, None),
        ("Payback Period (years)",   str(ret["payback_years"]) if ret["payback_years"] else "N/A", None, None),
    ]
    for lbl, val, fmt, fill_c in returns_rows:
        _label(ws, r, 1, lbl, bold=True)
        cell = ws.cell(r, 2, val)
        cell.font      = _font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _BORDER
        if fill_c:
            cell.fill = _fill(fill_c)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    _hdr_style(ws, r, 1, 2, "DEAL METRICS", _DARK_B); r += 1
    deal_metrics = [
        ("Entry Debt / EBITDA",       f"{tx['debt_ebitda']:.1f}x",           None),
        ("Exit Debt / EBITDA",        f"{ret['exit_debt_ebitda']:.1f}x",      None),
        ("Debt Paydown ($M)",         f"${ret['debt_paydown']:,.0f}M",         None),
        ("Total Mgmt Fees ($M)",      f"${ret['total_mgmt_fees']:,.0f}M",      None),
        ("Carry ($M)",                f"${ret['carry']:,.0f}M",                None),
        ("Entry Multiple",            f"{a.entry_ev_ebitda:.1f}x EV/EBITDA",  None),
        ("Exit Multiple",             f"{a.exit_ev_ebitda:.1f}x EV/EBITDA",   None),
    ]
    for lbl, val, fmt in deal_metrics:
        _label(ws, r, 1, lbl, indent=2)
        _write(ws, r, 2, val, align="left")
        r += 1

    ws.sheet_view.showGridLines = False


# ── Tab 9: Sensitivity ────────────────────────────────────────────────────────

def _color_irr(v: float) -> str:
    if math.isnan(v) or v < 0:
        return _RED_BG
    if v >= 0.20:
        return _GRN_BG
    if v >= 0.15:
        return _YLW_BG
    return _RED_BG


def _color_moic(v: float) -> str:
    if math.isnan(v) or v < 0:
        return _RED_BG
    if v >= 2.5:
        return _GRN_BG
    if v >= 2.0:
        return _YLW_BG
    return _RED_BG


def _write_sens_table(ws, start_row: int, start_col: int,
                      title: str, row_label: str, col_label: str,
                      row_vals: list, col_vals: list, table: list[list[float]],
                      fmt_fn,      # fn(val)->str
                      color_fn,    # fn(val)->hex
                      center_rc: tuple[int, int]) -> int:
    """Write a sensitivity table. Returns next available row."""
    r = start_row
    _hdr_style(ws, r, start_col, start_col + len(col_vals), title, _DARK_B); r += 1

    # Axis labels
    ws.cell(r, start_col, col_label).font = _font(bold=True)
    ws.cell(r, start_col).alignment = Alignment(horizontal="center")
    ws.cell(r, start_col).fill = _fill(_LT_B)

    for ci, cv in enumerate(col_vals):
        lbl = f"{cv:.1f}x" if abs(cv) < 50 else f"{cv*100:.0f}%"
        c = ws.cell(r, start_col + 1 + ci, lbl)
        c.font = _font(bold=True)
        c.fill = _fill(_LT_B)
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER
    r += 1

    for ri, (rv, row_data) in enumerate(zip(row_vals, table)):
        rl = ws.cell(r, start_col, f"{rv:.1f}x" if abs(rv) < 50 else f"{rv*100:.0f}%")
        rl.font = _font(bold=True)
        rl.fill = _fill(_LT_B)
        rl.alignment = Alignment(horizontal="center")
        rl.border = _BORDER
        # Row axis label in first col, first data row
        if ri == 0:
            ws.cell(start_row + 1, start_col, col_label).font = _font(bold=True)

        for ci, val in enumerate(row_data):
            display = fmt_fn(val)
            cell = ws.cell(r, start_col + 1 + ci, display)
            is_center = (ri == center_rc[0] and ci == center_rc[1])
            cell.fill      = _fill(_MED_B if is_center else color_fn(val))
            cell.font      = _font(bold=is_center, color=_FC_FORMULA)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = _BORDER
        r += 1

    # Legend
    r += 1
    for color, label in [(_GRN_BG, "Target (IRR >20% / MOIC >2.5x)"),
                         (_YLW_BG, "Acceptable (IRR 15-20% / MOIC 2.0-2.5x)"),
                         (_RED_BG, "Below Hurdle (IRR <15% / MOIC <2.0x)"),
                         (_MED_B,  "Base Case")]:
        ws.cell(r, start_col).fill = _fill(color)
        ws.cell(r, start_col).border = _BORDER
        ws.cell(r, start_col + 1, label).font = _font(size=9, italic=True)
        ws.merge_cells(start_row=r, start_column=start_col + 1,
                       end_row=r, end_column=start_col + len(col_vals))
        r += 1
    return r + 1


def _build_sensitivity(ws, data: dict) -> None:
    sens = data["sensitivity"]
    a    = data["assumptions"]

    ws.column_dimensions["A"].width = 14
    for i in range(1, 10):
        ws.column_dimensions[get_column_letter(i + 1)].width = 14

    r = 1
    _hdr_style(ws, r, 1, 8, "SENSITIVITY ANALYSIS", _NAVY, 13); r += 2

    # ── Table 1: IRR — entry multiple (rows) vs exit multiple (cols) ──────────
    _hdr_style(ws, r, 1, 6, "TABLE 1: GROSS IRR SENSITIVITY", _DARK_B); r += 1
    _label(ws, r, 1, "Rows: Entry EV/EBITDA | Columns: Exit EV/EBITDA", indent=0)
    r += 1
    _write_sens_table(
        ws, r, 1,
        "Entry Multiple (rows) × Exit Multiple (cols)",
        "Exit Multiple →", "Entry Multiple ↓",
        sens["entry_range"], sens["exit_range"],
        sens["irr_table"],
        _irr_fmt, _color_irr,
        center_rc=(2, 2),   # base case is center (index 2,2 in 5x5)
    )
    r += len(sens["entry_range"]) + 8

    # ── Table 2: MOIC — revenue CAGR vs EBITDA margin expansion ──────────────
    _hdr_style(ws, r, 1, 8, "TABLE 2: GROSS MOIC SENSITIVITY", _DARK_B); r += 1
    _label(ws, r, 1, "Rows: Revenue CAGR | Columns: EBITDA Margin Expansion/Year", indent=0)
    r += 1

    def _moic_fmt(v):
        if math.isnan(v) or v <= 0:
            return "N/M"
        return f"{v:.2f}x"

    mexp_pct = [f"{v*100:.1f}bps" for v in sens["margin_exp_range"]]
    cagr_pct = sens["cagr_range"]
    _write_sens_table(
        ws, r, 1,
        "Rev CAGR (rows) × Margin Expansion (cols)",
        "Margin Exp →", "CAGR ↓",
        cagr_pct, sens["margin_exp_range"],
        sens["moic_table2"],
        _moic_fmt, _color_moic,
        center_rc=(2, 2),
    )
    r += len(sens["cagr_range"]) + 8

    # ── Table 3: Leverage × Interest Rate ────────────────────────────────────
    _hdr_style(ws, r, 1, 5, "TABLE 3: LEVERAGE & RATE SENSITIVITY (GROSS IRR)", _DARK_B); r += 1
    _label(ws, r, 1, "Rows: Entry Debt/EBITDA | Columns: Rate Environment", indent=0)
    r += 1
    rate_labels = [f"{b*100:+.0f}bps" for b in sens["rate_bumps"]]
    lev_labels  = [f"{lv:.0f}x D/E" for lv in sens["lev_levels"]]
    _write_sens_table(
        ws, r, 1,
        "Debt/EBITDA (rows) × Rate Shift (cols)",
        "Rate Shift →", "Debt/EBITDA ↓",
        sens["lev_levels"], sens["rate_bumps"],
        sens["lev_table"],
        _irr_fmt, _color_irr,
        center_rc=(1, 1),   # 3x3 center is (1,1)
    )

    ws.sheet_view.showGridLines = False


# ── Main builder ──────────────────────────────────────────────────────────────

def build_lbo_excel(data: dict, output_path: str) -> None:
    wb = Workbook()

    tabs = [
        ("Cover",              _build_cover),
        ("Assumptions",        _build_assumptions),
        ("Transaction",        _build_transaction),
        ("Income Statement",   _build_income_statement),
        ("Balance Sheet",      _build_balance_sheet),
        ("Cash Flow",          _build_cash_flow),
        ("Debt Schedule",      _build_debt_schedule),
        ("Returns",            _build_returns),
        ("Sensitivity",        _build_sensitivity),
    ]

    # Use the default sheet for the first tab
    ws0 = wb.active
    ws0.title = tabs[0][0]
    tabs[0][1](ws0, data)

    for name, builder in tabs[1:]:
        ws = wb.create_sheet(name)
        builder(ws, data)

    # Tab color coding
    colors = ["1F4E79", "2E75B6", "2E75B6", "375623", "375623",
              "375623", "7030A0", "C00000", "833C00"]
    for ws, color in zip(wb.worksheets, colors):
        ws.sheet_properties.tabColor = color

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
