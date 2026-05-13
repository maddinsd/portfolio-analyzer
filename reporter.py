from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)

MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 600
SYSTEM_PROMPT = (
    "You are a financial analyst. Given this portfolio data, write a concise analysis "
    "covering: (1) overall risk profile, (2) sector concentration, (3) top 2 concerns, "
    "(4) one suggested improvement. Be direct. No disclaimers."
)


def build_payload(stats: dict) -> str:
    return json.dumps(stats, separators=(",", ":"))


def call_claude(payload: str) -> str:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set.")

    client = Anthropic(api_key=api_key)
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        system=[
            {
                "type": "text",
                "text": SYSTEM_PROMPT,
                "cache_control": {"type": "ephemeral"},
            }
        ],
        messages=[{"role": "user", "content": payload}],
    )
    return "".join(block.text for block in response.content if block.type == "text").strip()


def _fmt_money(value) -> str:
    if value is None:
        return "—"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    for suffix, threshold in (("T", 1e12), ("B", 1e9), ("M", 1e6)):
        if abs(v) >= threshold:
            return f"${v / threshold:.2f}{suffix}"
    return f"${v:,.0f}"


def _fmt(value, suffix: str = "") -> str:
    if value is None:
        return "—"
    return f"{value}{suffix}"


def _holdings_table(holdings: list[dict]) -> str:
    header = (
        "| Ticker | Weight | Sector | Price | Market Cap | P/E | 52W Low | 52W High | 6mo Return | Daily Vol |\n"
        "|---|---|---|---|---|---|---|---|---|---|\n"
    )
    rows = []
    for h in holdings:
        rows.append(
            "| {ticker} | {weight:.1%} | {sector} | {price} | {mcap} | {pe} | {low} | {high} | {ret} | {vol} |".format(
                ticker=h["ticker"],
                weight=h["weight"],
                sector=h["sector"] or "—",
                price=_fmt(h["current_price"]),
                mcap=_fmt_money(h["market_cap"]),
                pe=_fmt(h["trailing_pe"]),
                low=_fmt(h["fifty_two_week_low"]),
                high=_fmt(h["fifty_two_week_high"]),
                ret=_fmt(h["six_mo_return_pct"], "%"),
                vol=_fmt(h["daily_volatility_pct"], "%"),
            )
        )
    return header + "\n".join(rows)


def _correlation_table(corr: dict) -> str:
    tickers = list(corr.keys())
    header = "| | " + " | ".join(tickers) + " |\n"
    sep = "|" + "---|" * (len(tickers) + 1) + "\n"
    rows = []
    for row in tickers:
        cells = [f"{corr[row][col]:.3f}" if corr[row][col] is not None else "—" for col in tickers]
        rows.append(f"| **{row}** | " + " | ".join(cells) + " |")
    return header + sep + "\n".join(rows)


def _sector_table(sector_weights: dict) -> str:
    header = "| Sector | Weight |\n|---|---|\n"
    rows = [f"| {sector} | {weight:.1%} |" for sector, weight in sector_weights.items()]
    return header + "\n".join(rows)


def assemble_report(stats: dict, analysis: str) -> str:
    portfolio = stats["portfolio"]
    date_str = datetime.now().strftime("%Y-%m-%d")
    parts = [
        f"# Portfolio Report — {date_str}",
        "",
        "## Holdings",
        f"**Tickers:** {', '.join(portfolio['tickers'])}  ",
        f"**Equal weight per holding:** {portfolio['equal_weight']:.1%}  ",
        f"**Avg 6mo return:** {portfolio['avg_six_mo_return_pct']}%  ",
        f"**Avg daily volatility:** {portfolio['avg_daily_volatility_pct']}%",
        "",
        _holdings_table(stats["holdings"]),
        "",
        "## Sector Concentration",
        "",
        _sector_table(portfolio["sector_weights"]),
        "",
        "## Correlation Matrix (daily returns, 6mo)",
        "",
        _correlation_table(stats["correlation"]),
        "",
        "## Analyst Commentary",
        "",
        analysis,
        "",
    ]
    return "\n".join(parts)


def write_report(content: str, reports_dir: Path) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    filename = f"portfolio_{datetime.now().strftime('%Y%m%d')}.md"
    path = reports_dir / filename
    path.write_text(content, encoding="utf-8")
    return path


# ── Excel helpers ─────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", start_color="1F3864")   # dark navy
_ALT_FILL    = PatternFill("solid", start_color="EEF2F7")   # light blue-grey
_GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
_RED_FILL    = PatternFill("solid", start_color="FFC7CE")
_GREEN_FONT  = Font(name="Arial", color="276221", size=10)
_RED_FONT    = Font(name="Arial", color="9C0006", size=10)
_HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=10)
_BODY_FONT   = Font(name="Arial", size=10)
_TITLE_FONT  = Font(name="Arial", bold=True, size=13)
_THIN        = Side(style="thin", color="BFBFBF")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT        = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


def _apply_header(cell, value: str) -> None:
    cell.value = value
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _BORDER
    cell.alignment = _CENTER


def _apply_body(cell, value, alt: bool = False) -> None:
    cell.value = value
    cell.font = _BODY_FONT
    cell.fill = _ALT_FILL if alt else PatternFill()
    cell.border = _BORDER
    cell.alignment = _CENTER


def _sheet_holdings(wb: Workbook, holdings: list[dict], portfolio: dict) -> None:
    ws = wb.create_sheet("Holdings")

    # ── title row ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = f"Holdings — {datetime.now().strftime('%B %d, %Y')}"
    title.font = _TITLE_FONT
    title.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    # ── summary stats ──────────────────────────────────────────────────────
    ws["A2"] = "Tickers:"
    ws["A2"].font = Font(name="Arial", bold=True, size=10)
    ws["B2"] = ", ".join(portfolio["tickers"])
    ws["B2"].font = _BODY_FONT

    ws["A3"] = "Avg 6mo Return:"
    ws["A3"].font = Font(name="Arial", bold=True, size=10)
    ws["B3"] = portfolio["avg_six_mo_return_pct"] / 100 if portfolio["avg_six_mo_return_pct"] is not None else None
    ws["B3"].font = _BODY_FONT
    ws["B3"].number_format = "0.0%;-0.0%;-"

    ws["A4"] = "Avg Daily Volatility:"
    ws["A4"].font = Font(name="Arial", bold=True, size=10)
    ws["B4"] = portfolio["avg_daily_volatility_pct"] / 100 if portfolio["avg_daily_volatility_pct"] is not None else None
    ws["B4"].font = _BODY_FONT
    ws["B4"].number_format = "0.000%;-0.000%;-"

    # ── column headers (row 6) ─────────────────────────────────────────────
    headers = ["Ticker", "Weight", "Sector", "Price ($)", "Market Cap", "P/E", "52W Low", "52W High", "6mo Return", "Daily Vol"]
    col_widths = [10, 9, 18, 11, 14, 8, 10, 10, 13, 11]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _apply_header(ws.cell(row=6, column=col), h)
        ws.column_dimensions[ws.cell(row=6, column=col).column_letter].width = w
    ws.row_dimensions[6].height = 18

    # ── data rows ─────────────────────────────────────────────────────────
    for i, h in enumerate(holdings):
        row = 7 + i
        alt = i % 2 == 1
        ret = h["six_mo_return_pct"]

        values = [
            h["ticker"],
            h["weight"],
            h["sector"] or "—",
            h["current_price"],
            h["market_cap"],
            h["trailing_pe"],
            h["fifty_two_week_low"],
            h["fifty_two_week_high"],
            ret / 100 if ret is not None else None,
            h["daily_volatility_pct"] / 100 if h["daily_volatility_pct"] is not None else None,
        ]
        formats = [
            None, "0.0%", None, "$#,##0.00", "$#,##0,,\"B\"",
            "0.0x", "$#,##0.00", "$#,##0.00", "0.0%;-0.0%;-", "0.000%;-0.000%;-",
        ]

        for col, (val, fmt) in enumerate(zip(values, formats), start=1):
            cell = ws.cell(row=row, column=col)
            _apply_body(cell, val, alt)
            if fmt:
                cell.number_format = fmt

        # colour-code the 6mo Return cell (column 9)
        ret_cell = ws.cell(row=row, column=9)
        if ret is not None:
            if ret >= 0:
                ret_cell.fill = _GREEN_FILL
                ret_cell.font = _GREEN_FONT
            else:
                ret_cell.fill = _RED_FILL
                ret_cell.font = _RED_FONT

        ws.row_dimensions[row].height = 16

    # freeze panes below header
    ws.freeze_panes = "A7"


def _sheet_sector(wb: Workbook, sector_weights: dict) -> None:
    ws = wb.create_sheet("Sector Exposure")

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "Sector Exposure"
    title.font = _TITLE_FONT
    title.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    _apply_header(ws["A2"], "Sector")
    _apply_header(ws["B2"], "Weight")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12

    for i, (sector, weight) in enumerate(sector_weights.items()):
        row = 3 + i
        alt = i % 2 == 1
        _apply_body(ws.cell(row=row, column=1), sector, alt)
        wt_cell = ws.cell(row=row, column=2)
        _apply_body(wt_cell, weight, alt)
        wt_cell.number_format = "0.0%"
        ws.row_dimensions[row].height = 16

    # ── pie chart ─────────────────────────────────────────────────────────
    n = len(sector_weights)
    data_ref   = Reference(ws, min_col=2, min_row=2, max_row=2 + n)   # includes header
    labels_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + n)

    chart = PieChart()
    chart.title = "Sector Weights"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.style = 10
    chart.dataLabels = chart.dataLabels  # keep default; openpyxl sets labels via series
    chart.width  = 15
    chart.height = 12

    ws.add_chart(chart, "D2")


def _sheet_analysis(wb: Workbook, stats: dict, analysis: str) -> None:
    ws = wb.create_sheet("Analysis")

    portfolio = stats["portfolio"]

    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "Analyst Commentary"
    title.font = _TITLE_FONT
    title.alignment = _LEFT
    ws.row_dimensions[1].height = 24

    # key metrics
    metrics = [
        ("Portfolio", ", ".join(portfolio["tickers"])),
        ("Report Date", datetime.now().strftime("%B %d, %Y")),
        ("Avg 6mo Return", f"{portfolio['avg_six_mo_return_pct']}%"),
        ("Avg Daily Volatility", f"{portfolio['avg_daily_volatility_pct']}%"),
        ("Equal Weight", f"{portfolio['equal_weight']:.1%}"),
    ]
    for i, (label, value) in enumerate(metrics):
        row = 2 + i
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = Font(name="Arial", bold=True, size=10)
        val = ws.cell(row=row, column=2, value=value)
        val.font = _BODY_FONT

    # divider
    ws.row_dimensions[8].height = 6

    # commentary label
    ws["A9"].value = "Commentary"
    ws["A9"].font = Font(name="Arial", bold=True, size=11)

    # analysis text — strip markdown symbols for clean reading
    clean = (
        analysis
        .replace("**", "")
        .replace("##", "")
        .replace("# ", "")
        .replace("①", "1.")
        .replace("②", "2.")
        .replace("③", "3.")
        .replace("④", "4.")
    )

    ws.merge_cells("A10:C40")
    text_cell = ws["A10"]
    text_cell.value = clean
    text_cell.font = Font(name="Arial", size=10)
    text_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.row_dimensions[10].height = 400


def write_excel_report(stats: dict, analysis: str, reports_dir: Path) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    _sheet_holdings(wb, stats["holdings"], stats["portfolio"])
    _sheet_sector(wb, stats["portfolio"]["sector_weights"])
    _sheet_analysis(wb, stats, analysis)

    filename = f"portfolio_{datetime.now().strftime('%Y%m%d')}.xlsx"
    path = reports_dir / filename
    wb.save(str(path))
    return path
