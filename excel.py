from __future__ import annotations

import math
import os
import re
import zipfile
from datetime import date as _date_type

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Design system ─────────────────────────────────────────────────────────────
_F        = "Calibri"

_NAVY     = "003366"   # deep navy  – primary headers
_BLUE     = "1F4E79"   # mid blue   – section sub-headers
_STEEL    = "2D5F8A"   # steel blue – column headers
_WHITE    = "FFFFFF"
_ROW_ALT  = "EFF3F9"   # very light blue-grey – alternating row
_LBL_BG   = "E4EDF7"   # light blue – label cell background
_LBL_ALT  = "D6E4F0"   # slightly darker label bg for alt rows
_BORD_CLR = "C2CBD5"   # border colour
_GRN_BG   = "E2F0D9";  _GRN_FG = "375623"
_RED_BG   = "FFE0E0";  _RED_FG = "9C0006"
_SUBTITLE = "EAF0F7"   # subtitle bar

_THIN   = Side(style="thin", color=_BORD_CLR)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", start_color=c, end_color=c)


def _f(size=10, bold=False, color=None) -> Font:
    return Font(name=_F, size=size, bold=bold, color=color or "000000")


def _clean(text: str) -> str:
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    text = re.sub(r"\*(.*?)\*",     r"\1", text)
    return text.strip()


# ── Column auto-fit ───────────────────────────────────────────────────────────

def _auto_fit(ws, min_w: int = 9, max_w: int = 55) -> None:
    """Fit column widths to content, skipping merged-cell spans."""
    skip: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        for r, c in mr.cells:
            if (r, c) != (mr.min_row, mr.min_col):
                skip.add((r, c))
        if mr.max_col > mr.min_col:               # anchor of multi-col merge
            skip.add((mr.min_row, mr.min_col))

    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in skip or cell.value is None:
                continue
            L = len(str(cell.value))
            if cell.font and cell.font.bold:
                L = int(L * 1.12)
            widths[cell.column] = max(widths.get(cell.column, min_w), L + 3)

    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(max_w, w))


# ── Snapshot sheet ────────────────────────────────────────────────────────────

def _safe_fmt(val, decimals: int = 2, pct: bool = False, mult=None) -> str:
    if val is None or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))):
        return "N/A"
    if mult:
        val = val * mult
    return f"{val:.{decimals}f}%" if pct else f"{val:.{decimals}f}"


def _large_fmt(val) -> str:
    if val is None:
        return "N/A"
    try:
        v = float(val)
    except (TypeError, ValueError):
        return "N/A"
    for suffix, threshold in (("T", 1e12), ("B", 1e9), ("M", 1e6)):
        if abs(v) >= threshold:
            return f"${v / threshold:.2f}{suffix}"
    return f"${v:.2f}"


def _snap_lbl(cell, text: str, alt: bool = False) -> None:
    cell.value = text
    cell.font  = _f(10, bold=True)
    cell.fill  = _fill(_LBL_ALT if alt else _LBL_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = _BORDER


def _snap_val(cell, text: str, positive_green: bool = False, alt: bool = False) -> None:
    cell.value = text
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _BORDER
    if positive_green and isinstance(text, str) and text not in ("N/A", "—"):
        try:
            n = float(text.replace("%", "").replace("$", "").replace(",", "")
                      .replace("T", "").replace("B", "").replace("M", ""))
            if n > 0:
                cell.font = _f(10, color=_GRN_FG); cell.fill = _fill(_GRN_BG); return
            if n < 0:
                cell.font = _f(10, color=_RED_FG); cell.fill = _fill(_RED_BG); return
        except ValueError:
            pass
    cell.font = _f(10)
    cell.fill = _fill(_ROW_ALT if alt else _WHITE)


def _snap_section(ws, row: int, text: str, ncols: int = 4) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font  = _f(10, bold=True, color=_WHITE)
    c.fill  = _fill(_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _build_snapshot_sheet(wb, ticker: str, stats: dict) -> None:
    ws   = wb.create_sheet("Snapshot")
    info = stats["info"]
    f, fl = _safe_fmt, _large_fmt

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"{ticker}   ·   Stock Snapshot"
    t.font  = _f(14, bold=True, color=_WHITE)
    t.fill  = _fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Date bar
    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = f"As of {_date_type.today().strftime('%B %d, %Y')}"
    s.font  = _f(9, color="5A6B7B")
    s.fill  = _fill(_SUBTITLE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    sections = [
        ("PRICE & PERFORMANCE", True, [
            ("Current Price",   f"${stats['current_price']:.2f}",           "Prev Close",         f"${f(info.get('previousClose'))}"),
            ("6mo Return",      f(stats["stock_return_6mo"],  pct=True),    "S&P 500 6mo Return", f(stats["sp500_return_6mo"], pct=True)),
            ("Relative Return", f(stats["relative_return"],   pct=True),    "Annualized Vol",     f(stats["volatility_annualized"], pct=True)),
            ("52W High",        f"${f(info.get('fiftyTwoWeekHigh'))}",       "52W Low",            f"${f(info.get('fiftyTwoWeekLow'))}"),
            ("From 52W High",   f(stats["pct_from_52w_high"], pct=True),    "From 52W Low",       f(stats["pct_from_52w_low"], pct=True)),
            ("50-Day MA",       f"${f(info.get('fiftyDayAverage'))}",        "200-Day MA",         f"${f(info.get('twoHundredDayAverage'))}"),
        ]),
        ("VALUATION", False, [
            ("Trailing P/E",  f(info.get("trailingPE")),   "Forward P/E", f(info.get("forwardPE"))),
            ("Price / Book",  f(info.get("priceToBook")),  "Market Cap",  fl(info.get("marketCap"))),
            ("EPS (TTM)",     f(info.get("trailingEps")),  "EPS (Fwd)",   f(info.get("forwardEps"))),
        ]),
        ("FUNDAMENTALS", True, [
            ("Total Revenue",  fl(info.get("totalRevenue")),                             "Revenue Growth YoY", f(stats["revenue_growth_yoy"], pct=True)),
            ("Gross Margin",   f(info.get("grossMargins"),  mult=100, pct=True),         "Operating Margin",   f(info.get("operatingMargins"), mult=100, pct=True)),
            ("Net Margin",     f(info.get("profitMargins"), mult=100, pct=True),         "Free Cash Flow",     fl(info.get("freeCashflow"))),
            ("Total Debt",     fl(info.get("totalDebt")),                                "Dividend Yield",     f(info.get("dividendYield"), pct=True)),
        ]),
        ("ANALYST CONSENSUS", False, [
            ("Rating  (1 = Strong Buy)", f(info.get("recommendationMean")), "Price Target", f"${f(info.get('targetMeanPrice'))}"),
            ("# of Analysts",           str(info.get("numberOfAnalystOpinions") or "N/A"), "Sector",  info.get("sector") or "N/A"),
            ("Industry",                info.get("industry") or "N/A",                     "",        ""),
        ]),
    ]

    row = 3
    for section_name, use_color, rows_data in sections:
        ws.row_dimensions[row].height = 5          # thin spacer
        row += 1
        _snap_section(ws, row, section_name)
        row += 1
        alt = False
        for ll, lv, rl, rv in rows_data:
            _snap_lbl(ws.cell(row=row, column=1), ll, alt=alt)
            _snap_val(ws.cell(row=row, column=2), lv, positive_green=use_color, alt=alt)
            if rl:
                _snap_lbl(ws.cell(row=row, column=3), rl, alt=alt)
                _snap_val(ws.cell(row=row, column=4), rv, positive_green=use_color, alt=alt)
            ws.row_dimensions[row].height = 18
            row += 1
            alt = not alt

    _auto_fit(ws, min_w=14)


# ── Price Chart sheet ─────────────────────────────────────────────────────────

def _build_chart_sheet(wb, ticker: str, price_history, sp500_history) -> None:
    ws = wb.create_sheet("Price Chart")

    stock_close = price_history["Close"].reset_index()
    sp_close    = sp500_history["Close"].reset_index()
    stock_close.columns = ["Date", "Close"]
    sp_close.columns    = ["Date", "Close"]

    stock_idx = (stock_close["Close"] / stock_close["Close"].iloc[0] * 100).round(2)
    sp_idx    = (sp_close["Close"]    / sp_close["Close"].iloc[0]    * 100).round(2)

    # Column headers
    for ci, label in enumerate(["Date", ticker, "S&P 500"], start=1):
        c = ws.cell(row=1, column=ci)
        c.value = label
        c.font  = _f(10, bold=True, color=_WHITE)
        c.fill  = _fill(_NAVY)
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
        c.border = _BORDER
    ws.row_dimensions[1].height = 22

    n = len(stock_idx)
    for i in range(n):
        d = stock_close["Date"].iloc[i]
        try:
            lbl = d.strftime("%b '%y") if hasattr(d, "strftime") else str(d)[:10]
        except Exception:
            lbl = str(d)[:10]

        ws.cell(row=i + 2, column=1).value = lbl
        ws.cell(row=i + 2, column=2).value = float(stock_idx.iloc[i])
        if i < len(sp_idx):
            ws.cell(row=i + 2, column=3).value = float(sp_idx.iloc[i])

    all_v = list(stock_idx) + list(sp_idx)
    pad   = (max(all_v) - min(all_v)) * 0.08
    y_min = round(max(0.0, min(all_v) - pad), 1)
    y_max = round(max(all_v) + pad, 1)

    chart = LineChart()
    chart.title         = f"{ticker} vs. S&P 500  —  6-Month Relative Performance (Indexed to 100)"
    chart.style         = 10
    chart.y_axis.title  = "Indexed Value  (Base = 100)"
    chart.x_axis.title  = "Date"
    chart.width         = 32
    chart.height        = 18
    chart.y_axis.numFmt = "0.0"
    chart.y_axis.scaling.min = y_min
    chart.y_axis.scaling.max = y_max
    chart.x_axis.tickLblSkip = 21          # ~one label per month

    # Explicit y-axis gridlines
    try:
        from openpyxl.chart.axis import ChartLines
        chart.y_axis.majorGridlines = ChartLines()
    except Exception:
        pass

    # Legend below the plot
    legend          = Legend()
    legend.position = "b"
    legend.overlay  = False
    chart.legend    = legend

    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=n + 1), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=n + 1), titles_from_data=True)

    for s in chart.series:
        s.smooth = True

    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    ws.add_chart(chart, "E2")

    _auto_fit(ws, min_w=8, max_w=16)


# ── Analysis sheet ────────────────────────────────────────────────────────────

def _build_analysis_sheet(wb, markdown: str) -> None:
    ws    = wb.create_sheet("Analysis")
    lines = markdown.split("\n")

    # Pre-scan table column widths
    tbl_max = [0, 0, 0]
    for line in lines:
        if not line.startswith("|") or re.match(r"^\|[\s\-:|]+\|$", line):
            continue
        cols = [_clean(c) for c in line.strip("|").split("|")]
        for idx, val in enumerate(cols[:3]):
            tbl_max[idx] = max(tbl_max[idx], len(val))

    col_b = max(min(max(tbl_max[0], 30) + 4, 58), 42)
    col_c = min(tbl_max[1] + 4, 32) if tbl_max[1] else 26
    col_d = min(tbl_max[2] + 4, 28) if tbl_max[2] else 20
    text_w = col_b + col_c + col_d

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = col_b
    ws.column_dimensions["C"].width = col_c
    ws.column_dimensions["D"].width = col_d

    row   = 1
    i     = 0
    first = True

    while i < len(lines):
        line = lines[i]

        # H1 — report title
        if line.startswith("# "):
            ws.merge_cells(f"A{row}:D{row}")
            c = ws.cell(row=row, column=1)
            c.value = _clean(line[2:])
            c.font  = _f(14, bold=True, color=_WHITE)
            c.fill  = _fill(_NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 40
            row += 1
            ws.row_dimensions[row].height = 14   # spacer
            row += 1
            i += 1

        # H2 — section header
        elif line.startswith("## "):
            if not first:
                ws.row_dimensions[row].height = 10
                row += 1
            first = False
            ws.cell(row=row, column=1).fill = _fill(_BLUE)   # left accent sliver
            ws.merge_cells(f"B{row}:D{row}")
            c = ws.cell(row=row, column=2)
            c.value = _clean(line[3:]).upper()
            c.font  = _f(11, bold=True, color=_WHITE)
            c.fill  = _fill(_BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 26
            row += 1
            i += 1

        # Table block
        elif line.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].startswith("|"):
                table_lines.append(lines[i])
                i += 1

            is_hdr, alt = True, False
            for tl in table_lines:
                if re.match(r"^\|[\s\-:|]+\|$", tl):
                    continue
                cols  = [_clean(c) for c in tl.strip("|").split("|")]
                widths = [col_b, col_c, col_d]
                row_h  = 18
                for ci, val in enumerate(cols[:3]):
                    row_h = max(row_h, math.ceil(len(val) / max(1, widths[ci])) * 14 + 4)

                for ci, val in enumerate(cols[:3]):
                    c = ws.cell(row=row, column=ci + 2)
                    c.value  = val
                    c.border = _BORDER
                    if is_hdr:
                        c.font  = _f(9, bold=True, color=_WHITE)
                        c.fill  = _fill(_STEEL)
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.font  = _f(9)
                        c.fill  = _fill(_ROW_ALT if alt else _WHITE)
                        c.alignment = Alignment(horizontal="left", vertical="center",
                                                wrap_text=True, indent=1)
                ws.row_dimensions[row].height = row_h
                row    += 1
                alt     = not alt
                is_hdr  = False

            ws.row_dimensions[row].height = 5
            row += 1

        # Bullet
        elif line.startswith("- ") or line.startswith("* "):
            ws.merge_cells(f"B{row}:D{row}")
            c     = ws.cell(row=row, column=2)
            text  = "  •  " + _clean(line[2:])
            c.value = text
            c.font  = _f(11)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            n_lines = max(1, math.ceil(len(text) / text_w))
            ws.row_dimensions[row].height = max(18, n_lines * 16)
            row += 1
            i   += 1

        # Blank / HR
        elif line.strip() in ("---", "***", ""):
            ws.row_dimensions[row].height = 6
            row += 1
            i   += 1

        # Paragraph
        else:
            ws.merge_cells(f"B{row}:D{row}")
            c    = ws.cell(row=row, column=2)
            text = _clean(line)
            c.value = text
            c.font  = _f(11)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            n_lines = max(1, math.ceil(len(text) / text_w))
            ws.row_dimensions[row].height = max(16, n_lines * 16)
            row += 1
            i   += 1


# ── Financial statement sheet helpers ────────────────────────────────────────

def _fmt_m(val_m) -> str:
    if val_m is None:
        return "N/A"
    try:
        v = float(val_m)
    except (TypeError, ValueError):
        return "N/A"
    return f"${v / 1000:.1f}B" if abs(v) >= 1000 else f"${v:.0f}M"


def _fmt_pct_fin(val, signed: bool = False) -> str:
    if val is None:
        return "N/A"
    try:
        v = float(val)
        return f"{v:+.1f}%" if signed else f"{v:.1f}%"
    except (TypeError, ValueError):
        return "N/A"


def _fin_title(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font  = _f(13, bold=True, color=_WHITE)
    c.fill  = _fill(_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32


def _fin_subhdr(ws, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font  = _f(10, bold=True, color=_WHITE)
    c.fill  = _fill(_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def _fin_col_hdr(ws, row: int, headers: list[str]) -> None:
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=ci + 1)
        c.value = h
        c.font  = _f(9, bold=True, color=_WHITE)
        c.fill  = _fill(_STEEL)
        c.alignment = Alignment(horizontal="left" if ci == 0 else "right", vertical="center")
        c.border = _BORDER
    ws.row_dimensions[row].height = 18


def _fin_data_row(ws, row: int, label: str, values: list[str],
                  alt: bool = False, is_growth: bool = False) -> None:
    lbl_bg  = _LBL_ALT if alt else _LBL_BG
    data_bg = _ROW_ALT if alt else _WHITE

    lbl = ws.cell(row=row, column=1)
    lbl.value = label
    lbl.font  = _f(9, bold=True)
    lbl.fill  = _fill(lbl_bg)
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl.border = _BORDER

    for ci, val in enumerate(values):
        c = ws.cell(row=row, column=ci + 2)
        c.value  = val
        c.font   = _f(9)
        c.fill   = _fill(data_bg)
        c.border = _BORDER
        c.alignment = Alignment(horizontal="right", vertical="center")
        if is_growth and val not in ("N/A", "", None):
            try:
                n = float(str(val).replace("%", "").replace("+", ""))
                if n > 0:
                    c.fill = _fill(_GRN_BG); c.font = _f(9, color=_GRN_FG)
                elif n < 0:
                    c.fill = _fill(_RED_BG); c.font = _f(9, color=_RED_FG)
            except ValueError:
                pass
    ws.row_dimensions[row].height = 17


def _draw_table(ws, start_row: int, subtitle: str, periods: list[str],
                data_rows: list[tuple], ncols: int) -> int:
    _fin_subhdr(ws, start_row, subtitle, ncols)
    start_row += 1
    _fin_col_hdr(ws, start_row, [""] + periods)
    start_row += 1
    alt = False
    for label, values, is_growth in data_rows:
        _fin_data_row(ws, start_row, label, values, alt=alt, is_growth=is_growth)
        start_row += 1
        alt = not alt
    return start_row + 2   # blank gap


# ── Sheet 4: Income Statement ─────────────────────────────────────────────────

def _build_income_sheet(wb, fin_data: dict) -> None:
    ws  = wb.create_sheet("Income Statement")
    inc = fin_data.get("income_statement", {})
    q   = inc.get("quarterly")
    a   = inc.get("annual")
    N   = 5   # label col + 4 period cols

    _fin_title(ws, 1, "INCOME STATEMENT", N)
    row = 3

    def _rows(data, annual=False):
        if not data:
            return []
        r = [
            ("Revenue",           [_fmt_m(v) for v in data["revenue"]],           False),
            ("Gross Profit",      [_fmt_m(v) for v in data["gross_profit"]],      False),
            ("Operating Income",  [_fmt_m(v) for v in data["operating_income"]],  False),
            ("Net Income",        [_fmt_m(v) for v in data["net_income"]],        False),
            ("Gross Margin",      [_fmt_pct_fin(v) for v in data["gross_margin"]],     False),
            ("Operating Margin",  [_fmt_pct_fin(v) for v in data["operating_margin"]], False),
            ("Net Margin",        [_fmt_pct_fin(v) for v in data["net_margin"]],       False),
        ]
        if annual:
            if data.get("yoy_revenue"):
                r.append(("YoY Revenue Growth",
                           [_fmt_pct_fin(v, signed=True) for v in data["yoy_revenue"]], True))
            if data.get("yoy_ni"):
                r.append(("YoY Net Income Growth",
                           [_fmt_pct_fin(v, signed=True) for v in data["yoy_ni"]], True))
        return r

    if q:
        row = _draw_table(ws, row, "Quarterly  —  Last 4 Quarters", q["dates"], _rows(q), N)
    if a:
        row = _draw_table(ws, row, "Annual  —  Last 4 Years", a["dates"], _rows(a, annual=True), N)

    _auto_fit(ws, min_w=10)


# ── Sheet 5: Balance Sheet ────────────────────────────────────────────────────

def _build_balance_sheet_sheet(wb, fin_data: dict) -> None:
    ws = wb.create_sheet("Balance Sheet")
    bs = fin_data.get("balance_sheet") or {}
    dt = bs.get("date", "")

    _fin_title(ws, 1, f"BALANCE SHEET  —  {dt}" if dt else "BALANCE SHEET", 2)
    row = 3

    def _section(title, items, start):
        _fin_subhdr(ws, start, title, 2)
        start += 1
        _fin_col_hdr(ws, start, ["Metric", "Value"])
        start += 1
        alt = False
        for label, val in items:
            _fin_data_row(ws, start, label, [val], alt=alt)
            start += 1
            alt = not alt
        return start + 2

    row = _section("Assets, Liabilities & Equity", [
        ("Total Assets",         _fmt_m(bs.get("total_assets"))),
        ("Total Liabilities",    _fmt_m(bs.get("total_liabilities"))),
        ("Shareholders' Equity", _fmt_m(bs.get("shareholders_equity"))),
        ("Cash & Equivalents",   _fmt_m(bs.get("cash"))),
        ("Total Debt",           _fmt_m(bs.get("total_debt"))),
        ("Net Debt",             _fmt_m(bs.get("net_debt"))),
    ], row)

    row = _section("Key Ratios", [
        ("Current Ratio",  f"{bs['current_ratio']:.2f}"  if bs.get("current_ratio")  else "N/A"),
        ("Debt / Equity",  f"{bs['debt_to_equity']:.2f}" if bs.get("debt_to_equity") else "N/A"),
    ], row)

    _auto_fit(ws, min_w=10)


# ── Sheet 6: Cash Flow ────────────────────────────────────────────────────────

def _build_cashflow_sheet(wb, fin_data: dict) -> None:
    ws = wb.create_sheet("Cash Flow")
    cf = fin_data.get("cash_flow", {})
    q  = cf.get("quarterly")
    a  = cf.get("annual")
    N  = 5

    _fin_title(ws, 1, "CASH FLOW STATEMENT", N)
    row = 3

    def _rows(data, annual=False):
        if not data:
            return []
        r = [
            ("Operating Cash Flow", [_fmt_m(v) for v in data["operating_cash_flow"]], False),
            ("Capital Expenditure",  [_fmt_m(v) for v in data["capital_expenditure"]], False),
            ("Free Cash Flow",       [_fmt_m(v) for v in data["free_cash_flow"]],      False),
            ("FCF Margin",           [_fmt_pct_fin(v) for v in data["fcf_margin"]],    False),
        ]
        if annual and data.get("yoy_fcf"):
            r.append(("YoY FCF Growth",
                       [_fmt_pct_fin(v, signed=True) for v in data["yoy_fcf"]], True))
        return r

    if q:
        row = _draw_table(ws, row, "Quarterly  —  Last 4 Quarters", q["dates"], _rows(q), N)
    if a:
        row = _draw_table(ws, row, "Annual  —  Last 4 Years", a["dates"], _rows(a, annual=True), N)

    _auto_fit(ws, min_w=10)


# ── Sheet 4: Bull vs Bear ────────────────────────────────────────────────────

_BULL_DARK   = "1E6630"   # dark forest green – bull header / text
_BEAR_DARK   = "C00000"   # dark crimson      – bear header / text
_BULL_BG_ALT = "D4EBC4"   # slightly deeper green for alt rows
_BEAR_BG_ALT = "FFCECE"   # slightly deeper red   for alt rows


def _parse_section(markdown: str, header: str) -> list[str]:
    """Return non-empty lines from a ## header block, stripped of markdown markers."""
    m = re.search(
        rf"^## {re.escape(header)}\s*\n([\s\S]*?)(?=\n## |\Z)",
        markdown, re.MULTILINE,
    )
    if not m:
        return []
    lines = []
    for line in m.group(1).split("\n"):
        text = re.sub(r"^[-*•\d]+[.)]\s*", "", line.strip())
        text = _clean(text)
        if text:
            lines.append(text)
    return lines


def _build_bull_bear_sheet(wb, markdown: str) -> None:
    ws   = wb.create_sheet("Bull vs Bear")
    bull = _parse_section(markdown, "Bull Case")
    bear = _parse_section(markdown, "Bear Case")

    if not bull and not bear:
        return

    # Title
    ws.merge_cells("A1:B1")
    t           = ws["A1"]
    t.value     = "BULL CASE  ·  BEAR CASE"
    t.font      = _f(14, bold=True, color=_WHITE)
    t.fill      = _fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Column headers
    for col, label, bg in ((1, "BULL CASE", _BULL_DARK), (2, "BEAR CASE", _BEAR_DARK)):
        c           = ws.cell(row=2, column=col)
        c.value     = label
        c.font      = _f(12, bold=True, color=_WHITE)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _BORDER
    ws.row_dimensions[2].height = 28

    # Content rows
    COL_W = 58
    for i in range(max(len(bull), len(bear))):
        row    = i + 3
        alt    = i % 2 == 1
        b_text = ("•  " + bull[i]) if i < len(bull) else ""
        r_text = ("•  " + bear[i]) if i < len(bear) else ""

        b           = ws.cell(row=row, column=1)
        b.value     = b_text
        b.font      = _f(10, color=_BULL_DARK if b_text else "000000")
        b.fill      = _fill(_BULL_BG_ALT if alt else _GRN_BG)
        b.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True, indent=1)
        b.border    = _BORDER

        r           = ws.cell(row=row, column=2)
        r.value     = r_text
        r.font      = _f(10, color=_BEAR_DARK if r_text else "000000")
        r.fill      = _fill(_BEAR_BG_ALT if alt else _RED_BG)
        r.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True, indent=1)
        r.border    = _BORDER

        n_lines = max(1, math.ceil(max(len(b_text), len(r_text)) / COL_W))
        ws.row_dimensions[row].height = max(26, n_lines * 16 + 6)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60


# ── Sheet 7 (now 8): News & Sentiment ────────────────────────────────────────

_SENT_ROW: dict[str, tuple[str, str]] = {
    "Positive": (_GRN_BG, _GRN_FG),
    "Negative": (_RED_BG, _RED_FG),
    "Neutral":  ("F2F2F2", "595959"),
}
_OVERALL_ROW: dict[str, tuple[str, str]] = {
    "Bullish": (_GRN_BG, _GRN_FG),
    "Bearish": (_RED_BG, _RED_FG),
    "Neutral": ("F2F2F2", "595959"),
}


def _build_news_sheet(wb, sent_data: dict | None) -> None:
    if not sent_data:
        return

    ws       = wb.create_sheet("News & Sentiment")
    overall  = sent_data.get("overall", "Neutral")
    score    = sent_data.get("score")
    momentum = sent_data.get("momentum", "Stable")
    themes   = sent_data.get("themes", [])
    catalyst = sent_data.get("catalyst", "")
    summary  = sent_data.get("summary", "")
    articles = sent_data.get("articles", [])

    NCOLS    = 7
    last_col = get_column_letter(NCOLS)

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    t           = ws["A1"]
    t.value     = "NEWS & SENTIMENT ANALYSIS"
    t.font      = _f(14, bold=True, color=_WHITE)
    t.fill      = _fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Overall sentiment banner ────────────────────────────────────────────
    ov_bg, ov_fg = _OVERALL_ROW.get(overall, _OVERALL_ROW["Neutral"])
    ws.merge_cells(f"A2:{last_col}2")
    ov           = ws["A2"]
    score_part   = f"   ·   Score: {score:.1f}/10" if score is not None else ""
    moment_part  = f"   ·   Momentum: {momentum}" if momentum else ""
    ov.value     = f"Overall Sentiment: {overall}{score_part}{moment_part}"
    ov.font      = _f(13, bold=True, color=ov_fg)
    ov.fill      = _fill(ov_bg)
    ov.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    row = 3

    # ── Analyst summary ─────────────────────────────────────────────────────
    if summary:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = summary
        c.font      = _f(10, color="2C3E50")
        c.fill      = _fill(_SUBTITLE)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=2)
        n_lines = max(2, math.ceil(len(summary) / 130))
        ws.row_dimensions[row].height = max(40, n_lines * 16 + 8)
        row += 1

    # ── Key Themes ──────────────────────────────────────────────────────────
    if themes:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = "KEY THEMES   ·   " + "     ·     ".join(themes)
        c.font      = _f(10, bold=True, color=_NAVY)
        c.fill      = _fill("DCE9F7")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border    = Border(bottom=_THIN)
        ws.row_dimensions[row].height = 22
        row += 1

    # ── Catalyst Watch ──────────────────────────────────────────────────────
    if catalyst:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = "CATALYST WATCH   ·   " + catalyst
        c.font      = _f(10, color="595959")
        c.fill      = _fill("F2F6FC")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border    = Border(bottom=_THIN)
        ws.row_dimensions[row].height = 22
        row += 1

    # Spacer
    ws.row_dimensions[row].height = 8
    row += 1

    # ── Column headers ──────────────────────────────────────────────────────
    headers = ["Impact", "Type", "Sentiment", "Source", "Date",
               "Headline", "Investment Implication"]
    for ci, h in enumerate(headers, start=1):
        c           = ws.cell(row=row, column=ci)
        c.value     = h
        c.font      = _f(9, bold=True, color=_WHITE)
        c.fill      = _fill(_STEEL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Article rows ────────────────────────────────────────────────────────
    _IMP_FG = {"Major": "7B0000", "Minor": "4A4A4A"}

    for article in articles:
        sentiment        = article.get("sentiment", "Neutral")
        impact           = article.get("impact", "Moderate")
        row_bg, row_fg   = _SENT_ROW.get(sentiment, _SENT_ROW["Neutral"])

        values = [
            impact,
            article.get("type", ""),
            sentiment,
            article.get("source", ""),
            article.get("date", ""),
            article.get("headline", ""),
            article.get("note", ""),
        ]
        for ci, val in enumerate(values, start=1):
            c        = ws.cell(row=row, column=ci)
            c.value  = val
            c.fill   = _fill(row_bg)
            c.border = _BORDER

            if ci == 1:   # Impact
                c.font      = _f(9, bold=True, color=_IMP_FG.get(impact, row_fg))
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == 3: # Sentiment
                c.font      = _f(9, bold=True, color=row_fg)
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == 6: # Headline — hyperlinked
                url = article.get("url", "")
                if url:
                    c.hyperlink = url
                    c.font      = _f(9, color="0563C1")
                else:
                    c.font      = _f(9)
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True, indent=1)
            elif ci == 7: # Investment Implication
                c.font      = _f(9, color="1A1A2E")
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True, indent=1)
            else:
                c.font      = _f(9)
                c.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row].height = 52
        row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    for col, w in zip("ABCDEFG", [11, 13, 12, 18, 12, 44, 52]):
        ws.column_dimensions[col].width = w


# ── Research pipeline sheets ─────────────────────────────────────────────────

def _build_thesis_sheet(wb, research: dict | None) -> None:
    if not research:
        return
    th = research.get("thesis", {})
    ws = wb.create_sheet("Investment Thesis")

    ws.merge_cells("A1:B1")
    t           = ws["A1"]
    t.value     = "INVESTMENT THESIS"
    t.font      = _f(14, bold=True, color=_WHITE)
    t.fill      = _fill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    if th.get("_placeholder"):
        ws.merge_cells("A3:B3")
        c       = ws.cell(row=3, column=1)
        c.value = f"Analysis unavailable: {th.get('_error', 'unknown')}"
        c.font  = _f(10, color="999999")
        ws.column_dimensions["A"].width = 70
        ws.column_dimensions["B"].width = 70
        return

    rating = th.get("rating", "N/A")
    r_bg, r_fg = {
        "Buy":  (_GRN_BG, _GRN_FG),
        "Sell": (_RED_BG, _RED_FG),
    }.get(rating, ("FFF3CD", "856404"))

    ws.merge_cells("A2:B2")
    rb           = ws["A2"]
    rb.value     = f"Rating: {rating}   ·   Target: {th.get('target','N/A')}"
    rb.font      = _f(11, bold=True, color=r_fg)
    rb.fill      = _fill(r_bg)
    rb.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    row = 3

    def _section(header, points, hdr_bg, pt_bg, pt_bg_alt, pt_fg):
        nonlocal row
        ws.row_dimensions[row].height = 8
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        h           = ws.cell(row=row, column=1)
        h.value     = header
        h.font      = _f(10, bold=True, color=_WHITE)
        h.fill      = _fill(hdr_bg)
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        for i, pt in enumerate(points):
            ws.merge_cells(f"A{row}:B{row}")
            c           = ws.cell(row=row, column=1)
            c.value     = f"  {i+1}.  {pt}"
            c.font      = _f(10, color=pt_fg)
            c.fill      = _fill(pt_bg_alt if i % 2 else pt_bg)
            c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            c.border    = _BORDER
            ws.row_dimensions[row].height = max(22, math.ceil(len(pt) / 100) * 16 + 4)
            row += 1

    if th.get("bull"):
        _section("BULL CASE", th["bull"], _BULL_DARK, _GRN_BG, _BULL_BG_ALT, _BULL_DARK)
    if th.get("bear"):
        _section("BEAR CASE", th["bear"], _BEAR_DARK, _RED_BG, _BEAR_BG_ALT, _BEAR_DARK)

    if th.get("catalysts"):
        ws.row_dimensions[row].height = 8; row += 1
        ws.merge_cells(f"A{row}:B{row}")
        h           = ws.cell(row=row, column=1)
        h.value     = "KEY CATALYSTS"
        h.font      = _f(10, bold=True, color=_WHITE)
        h.fill      = _fill(_BLUE)
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22; row += 1
        for i, cat in enumerate(th["catalysts"]):
            ws.merge_cells(f"A{row}:B{row}")
            c           = ws.cell(row=row, column=1)
            c.value     = f"  ◆  {cat}"
            c.font      = _f(10, color=_NAVY)
            c.fill      = _fill("C8DDF0" if i % 2 else "DCE9F7")
            c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            c.border    = _BORDER
            ws.row_dimensions[row].height = max(22, math.ceil(len(cat) / 100) * 16 + 4)
            row += 1

    if th.get("verdict"):
        ws.row_dimensions[row].height = 8; row += 1
        ws.merge_cells(f"A{row}:B{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = f"VERDICT  ·  {th['verdict']}"
        c.font      = _f(10, bold=True, color=_NAVY)
        c.fill      = _fill(_SUBTITLE)
        c.alignment = Alignment(wrap_text=True, vertical="center", indent=2)
        ws.row_dimensions[row].height = max(30, math.ceil(len(th["verdict"]) / 100) * 16 + 6)

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 70


def _build_comps_sheet(wb, research: dict | None) -> None:
    if not research:
        return
    co = research.get("comps", {})
    ws = wb.create_sheet("Comps Analysis")
    N  = 6

    _fin_title(ws, 1, "COMPARABLE COMPANIES ANALYSIS  (multiples approximate)", N)

    if co.get("_placeholder"):
        ws.merge_cells(f"A3:{get_column_letter(N)}3")
        c       = ws.cell(row=3, column=1)
        c.value = f"Analysis unavailable: {co.get('_error','unknown')}"
        c.font  = _f(10, color="999999")
        return

    row = 3
    if co.get("summary"):
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = co["summary"]
        c.font      = _f(10, color="2C3E50")
        c.fill      = _fill(_SUBTITLE)
        c.alignment = Alignment(wrap_text=True, vertical="center", indent=2)
        ws.row_dimensions[row].height = max(40, math.ceil(len(co["summary"]) / 130) * 16 + 8)
        row += 1

    if co.get("premium"):
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = f"vs. Peer Median EV/EBITDA:  {co['premium']}"
        c.font      = _f(10, bold=True, color=_NAVY)
        c.fill      = _fill("DCE9F7")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.row_dimensions[row].height = 8; row += 1

    _fin_col_hdr(ws, row, ["Company", "Ticker", "EV/EBITDA", "P/E (Fwd)", "EV/Revenue", "Note"])
    row += 1

    def _mx(v):
        try: return f"{float(v):.1f}x"
        except Exception: return str(v) if v is not None else "N/A"

    for i, comp in enumerate(co.get("comps", [])):
        alt    = i % 2 == 1
        bg     = _ROW_ALT if alt else _WHITE
        values = [comp.get("company",""), comp.get("ticker",""),
                  _mx(comp.get("ev_ebitda")), _mx(comp.get("pe_fwd")),
                  _mx(comp.get("ev_rev")),    comp.get("note","")]
        for ci, val in enumerate(values):
            c           = ws.cell(row=row, column=ci + 1)
            c.value     = val
            c.font      = _f(9)
            c.fill      = _fill(bg)
            c.border    = _BORDER
            left        = ci in (0, 5)
            c.alignment = Alignment(horizontal="left" if left else "center",
                                    vertical="center", wrap_text=(ci == 5),
                                    indent=1 if left else 0)
        ws.row_dimensions[row].height = 18
        row += 1

    for col, w in zip("ABCDEF", [24, 10, 13, 13, 13, 36]):
        ws.column_dimensions[col].width = w


def _build_earnings_sheet(wb, research: dict | None) -> None:
    if not research:
        return
    ep = research.get("earnings", {})
    ws = wb.create_sheet("Earnings Preview")
    N  = 6

    _fin_title(ws, 1, "EARNINGS PREVIEW", N)

    if ep.get("_placeholder"):
        ws.merge_cells(f"A3:{get_column_letter(N)}3")
        c       = ws.cell(row=3, column=1)
        c.value = f"Analysis unavailable: {ep.get('_error','unknown')}"
        c.font  = _f(10, color="999999")
        return

    row = 3
    meta = "   ·   ".join(x for x in [
        f"Next Earnings: {ep['next_earnings']}"       if ep.get("next_earnings")   else None,
        f"Consensus Rev: {ep['consensus_rev']}"       if ep.get("consensus_rev")   else None,
        f"Consensus EPS: {ep['consensus_eps']}"       if ep.get("consensus_eps")   else None,
        f"Options Implied Move: {ep['implied_move']}" if ep.get("implied_move")    else None,
    ] if x)
    if meta:
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = meta
        c.font      = _f(10, bold=True, color=_NAVY)
        c.fill      = _fill(_SUBTITLE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26
        row += 1

    if ep.get("watch"):
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = "KEY METRICS TO WATCH   ·   " + "     ·     ".join(ep["watch"])
        c.font      = _f(10, bold=True, color=_NAVY)
        c.fill      = _fill("DCE9F7")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.row_dimensions[row].height = 8; row += 1

    _fin_col_hdr(ws, row, ["Scenario", "Revenue", "EPS", "Implied Move", "Probability", "Trigger / Key Driver"])
    row += 1

    _SCEN = {"Bull": (_GRN_BG, _GRN_FG), "Base": ("FFF8E1", "856404"), "Bear": (_RED_BG, _RED_FG)}
    for s in ep.get("scenarios", []):
        name    = s.get("name", "")
        bg, fg  = _SCEN.get(name, ("F2F2F2", "595959"))
        values  = [name, s.get("rev",""), s.get("eps",""),
                   s.get("move",""), s.get("prob",""), s.get("trigger","")]
        for ci, val in enumerate(values):
            c           = ws.cell(row=row, column=ci + 1)
            c.value     = val
            c.fill      = _fill(bg)
            c.border    = _BORDER
            if ci == 0:
                c.font      = _f(10, bold=True, color=fg)
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == 5:
                c.font      = _f(9, color="1A1A2E")
                c.alignment = Alignment(horizontal="left", vertical="top",
                                        wrap_text=True, indent=1)
            else:
                c.font      = _f(9, bold=True, color=fg)
                c.alignment = Alignment(horizontal="center", vertical="top")
        ws.row_dimensions[row].height = max(36, math.ceil(len(s.get("trigger","")) / 35) * 16 + 6)
        row += 1

    for col, w in zip("ABCDEF", [14, 13, 10, 15, 14, 42]):
        ws.column_dimensions[col].width = w


# ── DCF Model sheet ───────────────────────────────────────────────────────────

def _build_dcf_sheet(wb, dcf_result: dict | None, ticker: str) -> None:
    if not dcf_result or dcf_result.get("error"):
        return

    ws  = wb.create_sheet("DCF Model")
    inp = dcf_result["inputs"]
    fc  = dcf_result["forecast"]
    val = dcf_result["valuation"]
    sen = dcf_result["sensitivity"]
    px  = val["current_price"] or 0

    NCOLS = 6   # A–F used throughout

    # Title
    _fin_title(ws, 1, f"DCF MODEL  —  {ticker}", NCOLS)

    # Subtitle
    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    sub           = ws["A2"]
    sub.value     = "Two-Stage Discounted Cash Flow  ·  FCFF-based  ·  Gordon Growth Terminal Value"
    sub.font      = _f(9, color="5A6B7B")
    sub.fill      = _fill(_SUBTITLE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4   # row 3 is a blank spacer (default height)

    # ── Section 1: Model Assumptions ─────────────────────────────────────────
    _fin_subhdr(ws, row, "MODEL ASSUMPTIONS", NCOLS)
    row += 1
    _fin_col_hdr(ws, row, ["Assumption", "Value"])
    row += 1

    assumptions = [
        ("Risk-Free Rate",           f"{inp['rf']:.2f}%",          False),
        ("Equity Risk Premium",      f"{inp['erp']:.2f}%",         False),
        ("Beta (5-yr Monthly)",      f"{inp['beta']:.3f}",         False),
        ("Cost of Equity (Ke)",      f"{inp['ke']:.2f}%",          False),
        ("Pre-Tax Cost of Debt",     f"{inp['kd_pretax']:.2f}%",   False),
        ("After-Tax Cost of Debt",   f"{inp['kd_aftertax']:.2f}%", False),
        ("Effective Tax Rate",       f"{inp['tax_rate']:.2f}%",    False),
        ("Equity Weight",            f"{inp['we']:.1f}%",          False),
        ("Debt Weight",              f"{inp['wd']:.1f}%",          False),
        ("WACC",                     f"{inp['wacc']:.2f}%",        False),  # idx 9 — highlighted
        ("Revenue CAGR (3-yr Hist)", f"{inp['rev_cagr']:.2f}%",   False),
        ("EBIT Margin (Hist. Avg)",  f"{inp['ebit_margin']:.2f}%", False),
        ("D&A as % of Revenue",      f"{inp['da_pct']:.2f}%",      False),
        ("CapEx as % of Revenue",    f"{inp['capex_pct']:.2f}%",   False),
        ("ΔWC as % of Revenue", f"{inp['nwc_pct']:.2f}%",     False),
        ("Terminal Growth Rate",     f"{inp['tg']:.2f}%",          False),
        ("Forecast Horizon",         "5 Years",                    False),
    ]

    for idx, (label, value, _) in enumerate(assumptions):
        alt      = idx % 2 == 1
        is_wacc  = (idx == 9)
        lbl_bg   = _BLUE if is_wacc else (_LBL_ALT if alt else _LBL_BG)
        val_bg   = _BLUE if is_wacc else (_ROW_ALT if alt else _WHITE)
        txt_clr  = _WHITE if is_wacc else "000000"

        lbl           = ws.cell(row=row, column=1)
        lbl.value     = label
        lbl.font      = _f(9, bold=True, color=txt_clr)
        lbl.fill      = _fill(lbl_bg)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.border    = _BORDER

        vc           = ws.cell(row=row, column=2)
        vc.value     = value
        vc.font      = _f(9, bold=is_wacc, color=txt_clr)
        vc.fill      = _fill(val_bg)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = _BORDER

        ws.row_dimensions[row].height = 17
        row += 1

    row += 1  # spacer

    # ── Section 2: 5-Year FCFF Forecast ──────────────────────────────────────
    years = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    _fin_subhdr(ws, row, "5-YEAR FCFF FORECAST  ($M)", NCOLS)
    row += 1
    _fin_col_hdr(ws, row, ["Metric"] + years)
    row += 1

    forecast_rows = [
        ("Revenue Growth",  [f"{g:.1f}%" for g in fc["growth_pct"]], True),
        ("Revenue ($M)",    [_fmt_m(v) for v in fc["revenue_m"]],    False),
        ("EBIT ($M)",       [_fmt_m(v) for v in fc["ebit_m"]],       False),
        ("NOPAT ($M)",      [_fmt_m(v) for v in fc["nopat_m"]],      False),
        ("D&A ($M)",        [_fmt_m(v) for v in fc["da_m"]],         False),
        ("CapEx ($M)",      [_fmt_m(v) for v in fc["capex_m"]],      False),
        ("ΔNWC ($M)",  [_fmt_m(v) for v in fc["dnwc_m"]],       False),
        ("FCFF ($M)",       [_fmt_m(v) for v in fc["fcff_m"]],       False),
    ]

    alt = False
    for label, values, is_growth in forecast_rows:
        _fin_data_row(ws, row, label, values, alt=alt, is_growth=is_growth)
        row += 1
        alt = not alt

    row += 1  # spacer

    # ── Section 3: Valuation Bridge ──────────────────────────────────────────
    _fin_subhdr(ws, row, "VALUATION BRIDGE", NCOLS)
    row += 1
    _fin_col_hdr(ws, row, ["Metric", "Value"])
    row += 1

    upside_str = (f"{val['upside_pct']:+.1f}%"
                  if val["upside_pct"] is not None else "N/A")
    iv_str = f"${val['intrinsic']:.2f}"
    px_str = f"${px:.2f}" if px else "N/A"

    valuation_rows = [
        ("PV of Explicit Forecast",      _fmt_m(val["pv_fcff_m"]),  False),
        ("Terminal Value (undiscounted)", _fmt_m(val["tv_m"]),       False),
        ("PV of Terminal Value",          _fmt_m(val["pv_tv_m"]),    False),
        ("Terminal Value % of EV",        f"{val['tv_pct']:.1f}%",  False),
        ("Enterprise Value",              _fmt_m(val["ev_m"]),       False),  # idx 4 — highlighted
        ("Less: Net Debt",                _fmt_m(val["net_debt_m"]), False),
        ("Equity Value",                  _fmt_m(val["eq_val_m"]),   False),
        ("Shares Outstanding",            f"{val['shares_m']:.2f}M", False),
        ("Intrinsic Value / Share",       iv_str,                    False),  # idx 8 — highlighted
        ("Current Price",                 px_str,                    False),
        ("Implied Upside / Downside",     upside_str,                True),   # green/red
    ]

    for idx, (label, value, is_growth) in enumerate(valuation_rows):
        alt     = idx % 2 == 1
        is_ev   = (idx == 4)
        is_iv   = (idx == 8)
        special = is_ev or is_iv

        lbl_bg  = _BLUE if special else (_LBL_ALT if alt else _LBL_BG)
        val_bg  = _BLUE if special else (_ROW_ALT if alt else _WHITE)
        txt_clr = _WHITE if special else "000000"

        lbl           = ws.cell(row=row, column=1)
        lbl.value     = label
        lbl.font      = _f(9, bold=True, color=txt_clr)
        lbl.fill      = _fill(lbl_bg)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.border    = _BORDER

        vc = ws.cell(row=row, column=2)
        vc.value  = value
        vc.border = _BORDER
        vc.alignment = Alignment(horizontal="right", vertical="center")

        if special:
            vc.font = _f(9, bold=True, color=txt_clr)
            vc.fill = _fill(val_bg)
        elif is_growth and value not in ("N/A", ""):
            try:
                n = float(value.replace("%", "").replace("+", ""))
                if n > 0:
                    vc.font = _f(9, color=_GRN_FG); vc.fill = _fill(_GRN_BG)
                elif n < 0:
                    vc.font = _f(9, color=_RED_FG); vc.fill = _fill(_RED_BG)
                else:
                    vc.font = _f(9); vc.fill = _fill(val_bg)
            except ValueError:
                vc.font = _f(9); vc.fill = _fill(val_bg)
        else:
            vc.font = _f(9)
            vc.fill = _fill(val_bg)

        ws.row_dimensions[row].height = 17
        row += 1

    row += 1  # spacer

    # ── Section 4: Sensitivity Table ─────────────────────────────────────────
    _fin_subhdr(ws, row, "SENSITIVITY: INTRINSIC VALUE PER SHARE  ($/share)", NCOLS)
    row += 1

    ws.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
    note           = ws.cell(row=row, column=1)
    note.value     = "Row = WACC  ·  Column = Terminal Growth Rate  ·  Base case highlighted in navy"
    note.font      = _f(9, color="595959")
    note.fill      = _fill("F2F6FC")
    note.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 16
    row += 1

    tg_hdrs = [f"{t:.1f}%" for t in sen["tg_range"]]
    _fin_col_hdr(ws, row, ["WACC \\ Term Growth"] + tg_hdrs)
    row += 1

    for ri, (w, row_vals) in enumerate(zip(sen["wacc_range"], sen["table"])):
        is_base_row = (ri == 2)

        lbl           = ws.cell(row=row, column=1)
        lbl.value     = f"{w:.2f}%"
        lbl.font      = _f(9, bold=True, color=_WHITE if is_base_row else "000000")
        lbl.fill      = _fill(_NAVY if is_base_row else (_LBL_ALT if ri % 2 else _LBL_BG))
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.border    = _BORDER

        for ci, iv in enumerate(row_vals):
            is_base = is_base_row and ci == 2
            c        = ws.cell(row=row, column=ci + 2)
            c.border  = _BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            if iv is None:
                c.value = "N/A"
                c.font  = _f(9, color="999999")
                c.fill  = _fill("F2F2F2")
            elif is_base:
                c.value = f"${iv:.2f}"
                c.font  = _f(9, bold=True, color=_WHITE)
                c.fill  = _fill(_NAVY)
            elif px and iv > px:
                c.value = f"${iv:.2f}"
                c.font  = _f(9, color=_GRN_FG)
                c.fill  = _fill(_GRN_BG)
            elif px and iv < px:
                c.value = f"${iv:.2f}"
                c.font  = _f(9, color=_RED_FG)
                c.fill  = _fill(_RED_BG)
            else:
                c.value = f"${iv:.2f}"
                c.font  = _f(9)
                c.fill  = _fill(_ROW_ALT if ri % 2 else _WHITE)

        ws.row_dimensions[row].height = 18
        row += 1

    # Warning banner if any
    if dcf_result.get("warnings"):
        row += 1
        ws.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
        w           = ws.cell(row=row, column=1)
        w.value     = "  ".join(dcf_result["warnings"])
        w.font      = _f(9, bold=True, color=_RED_FG)
        w.fill      = _fill(_RED_BG)
        w.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14


# ── Entry point ───────────────────────────────────────────────────────────────

def _build_competitive_sheet(wb, comp_result: dict | None) -> None:
    if not comp_result:
        return

    ws = wb.create_sheet("Competitive Analysis")
    N  = 10   # columns A–J

    if comp_result.get("error"):
        _fin_title(ws, 1, "COMPETITIVE ANALYSIS", N)
        ws.merge_cells(f"A3:{get_column_letter(N)}3")
        c       = ws.cell(row=3, column=1)
        c.value = f"Data unavailable: {comp_result['error']}"
        c.font  = _f(10, color="999999")
        return

    target  = comp_result["target"]
    peers   = comp_result["peers"]
    medians = comp_result["peer_medians"]
    ranks   = comp_result["rankings"]
    claude  = comp_result.get("claude") or {}
    ticker  = target["ticker"]

    # ── Title ────────────────────────────────────────────────────────────────
    _fin_title(ws, 1, f"COMPETITIVE ANALYSIS  —  {ticker}", N)

    src_label = comp_result.get("source", "").title()
    sub_text  = f"Peer universe: {src_label}  ·  {len(peers)} comparable companies identified"
    ws.merge_cells(f"A2:{get_column_letter(N)}2")
    sub           = ws["A2"]
    sub.value     = sub_text
    sub.font      = _f(9, color="5A6B7B")
    sub.fill      = _fill(_SUBTITLE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4   # row 3 is a blank spacer

    # ── Section 1: Competitive Positioning Summary ────────────────────────────
    _fin_subhdr(ws, row, "COMPETITIVE POSITIONING SUMMARY", N)
    row += 1

    def _kv(r, key, val, alt=False, navy=False, red=False):
        bg_k = _NAVY if navy else (_RED_BG if red else (_LBL_ALT if alt else _LBL_BG))
        bg_v = _NAVY if navy else (_RED_BG if red else (_ROW_ALT if alt else _WHITE))
        fg_k = _WHITE if navy else (_RED_FG if red else "000000")
        fg_v = _WHITE if navy else (_RED_FG if red else "000000")

        ws.merge_cells(f"A{r}:B{r}")
        k           = ws.cell(row=r, column=1)
        k.value     = key
        k.font      = _f(9, bold=True, color=fg_k)
        k.fill      = _fill(bg_k)
        k.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        k.border    = _BORDER

        ws.merge_cells(f"C{r}:{get_column_letter(N)}{r}")
        v           = ws.cell(row=r, column=3)
        v.value     = val
        v.font      = _f(9, color=fg_v)
        v.fill      = _fill(bg_v)
        v.alignment = Alignment(horizontal="left", vertical="center", indent=1,
                                wrap_text=red)  # wrap the risk row
        v.border    = _BORDER
        ws.row_dimensions[r].height = 32 if red else 18

    _kv(row, "Moat Type",            claude.get("moat_type", "—"), navy=True);   row += 1
    _kv(row, "Moat Strength",        claude.get("moat_strength", "—"), alt=False); row += 1
    _kv(row, "Competitive Position", claude.get("position", "—"), alt=True);      row += 1
    _kv(row, "Key Risk (24mo)",      claude.get("key_risk", "—"), red=True);      row += 1

    if comp_result.get("warning"):
        ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
        w           = ws.cell(row=row, column=1)
        w.value     = f"Note: {comp_result['warning']}"
        w.font      = _f(9, color="7D5A00")
        w.fill      = _fill("FFF3CC")
        w.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        w.border    = _BORDER
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # spacer

    # ── Section 2: Peer Comparison Table ─────────────────────────────────────
    _fin_subhdr(ws, row, "PEER COMPARISON TABLE", N)
    row += 1
    _fin_col_hdr(ws, row, ["Company", "Ticker", "Mkt Cap",
                            "Rev Growth", "Gross Margin", "Op Margin",
                            "ROE", "D/E", "Fwd P/E", "vs Median P/E"])
    row += 1

    _YLW_BG = "FFF9C4"; _YLW_FG = "7D5A00"
    _RANK_CLR = {
        "top": (_GRN_BG, _GRN_FG),
        "mid": (_YLW_BG, _YLW_FG),
        "bot": (_RED_BG, _RED_FG),
    }

    def _cap(v):
        if v is None: return "N/A"
        if abs(v) >= 1e12: return f"${v/1e12:.1f}T"
        if abs(v) >= 1e9:  return f"${v/1e9:.1f}B"
        return f"${v/1e6:.0f}M"

    def _pp(v):
        return "N/A" if v is None else f"{v:.1f}%"

    def _pf(v):
        return "N/A" if v is None else f"{v:.1f}x"

    def _pm(v):
        return "N/A" if v is None else f"{v:+.1f}%"

    # Metric columns: index → (data field, rank key or None)
    METRIC_COLS = [
        (3, "rev_growth",   "rev_growth"),
        (4, "gross_margin", "gross_margin"),
        (5, "op_margin",    "op_margin"),
        (6, "roe",          "roe"),
        (7, "de",           None),
        (8, "fpe",          None),
    ]
    fpe_prem = ranks.get("fpe_premium_pct")
    target_row_vals = [
        target.get("name", "")[:22], target.get("ticker", ""),
        _cap(target.get("mktcap")),
        _pp(target.get("rev_growth")), _pp(target.get("gross_margin")),
        _pp(target.get("op_margin")),  _pp(target.get("roe")),
        _pp(target.get("de")),         _pf(target.get("fpe")),
        _pm(fpe_prem),
    ]

    # Target row — metric cells color-coded by tercile, identity cols navy
    for ci, val in enumerate(target_row_vals):
        c           = ws.cell(row=row, column=ci + 1)
        c.value     = val
        c.border    = _BORDER
        c.alignment = Alignment(horizontal="left" if ci == 0 else "right", vertical="center")

        if ci < 2:   # Company, Ticker — navy
            c.font = _f(9, bold=True, color=_WHITE)
            c.fill = _fill(_NAVY)
        elif ci == 2:  # Mkt Cap — navy
            c.font = _f(9, bold=True, color=_WHITE)
            c.fill = _fill(_NAVY)
        elif ci == 9:  # vs Median P/E — navy
            c.font = _f(9, bold=True, color=_WHITE)
            c.fill = _fill(_NAVY)
        else:
            rank_key = {3: "rev_growth", 4: "gross_margin", 5: "op_margin",
                        6: "roe"}.get(ci)
            if rank_key:
                r_str   = ranks.get(rank_key, "mid")
                bg, fg  = _RANK_CLR.get(r_str, (_ROW_ALT, "000000"))
                c.font  = _f(9, bold=True, color=fg)
                c.fill  = _fill(bg)
            else:
                c.font = _f(9, bold=True, color=_WHITE)
                c.fill = _fill(_NAVY)

    ws.row_dimensions[row].height = 20
    row += 1

    # Peer rows, sorted by market cap descending
    for idx, p in enumerate(sorted(peers, key=lambda x: x.get("mktcap") or 0, reverse=True)):
        alt = idx % 2 == 1
        bg  = _ROW_ALT if alt else _WHITE
        vals = [
            p.get("name", "")[:22], p.get("ticker", ""),
            _cap(p.get("mktcap")),
            _pp(p.get("rev_growth")), _pp(p.get("gross_margin")),
            _pp(p.get("op_margin")),  _pp(p.get("roe")),
            _pp(p.get("de")),         _pf(p.get("fpe")),
            "—",
        ]
        for ci, val in enumerate(vals):
            c           = ws.cell(row=row, column=ci + 1)
            c.value     = val
            c.font      = _f(9)
            c.fill      = _fill(bg)
            c.border    = _BORDER
            c.alignment = Alignment(horizontal="left" if ci == 0 else "right",
                                    vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    # Peer median row
    med_vals = [
        "Peer Median", "—", "—",
        _pp(medians.get("rev_growth")), _pp(medians.get("gross_margin")),
        _pp(medians.get("op_margin")),  _pp(medians.get("roe")),
        "—", _pf(medians.get("fpe")), "—",
    ]
    for ci, val in enumerate(med_vals):
        c           = ws.cell(row=row, column=ci + 1)
        c.value     = val
        c.font      = _f(9, bold=True, color=_WHITE)
        c.fill      = _fill(_BLUE)
        c.border    = _BORDER
        c.alignment = Alignment(horizontal="left" if ci == 0 else "right",
                                vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    row += 1  # spacer

    # ── Section 3: Clustered Bar Chart ────────────────────────────────────────
    _fin_subhdr(ws, row, f"{ticker} vs PEER MEDIAN — KEY MARGIN COMPARISON", N)
    row += 1

    # Write chart data table (categories as rows, series as columns)
    chart_data_row = row
    hdr_data = ["Metric", ticker, "Peer Median"]
    chart_rows = [
        ["Gross Margin",
         target.get("gross_margin"), medians.get("gross_margin")],
        ["Operating Margin",
         target.get("op_margin"),   medians.get("op_margin")],
        ["ROE",
         target.get("roe"),         medians.get("roe")],
    ]

    # Header row
    for ci, val in enumerate(hdr_data):
        c           = ws.cell(row=row, column=ci + 1)
        c.value     = val
        c.font      = _f(9, bold=True, color=_WHITE)
        c.fill      = _fill(_STEEL)
        c.alignment = Alignment(horizontal="left" if ci == 0 else "right",
                                vertical="center")
        c.border    = _BORDER
    ws.row_dimensions[row].height = 18
    row += 1

    for i, (label, tval, mval) in enumerate(chart_rows):
        alt = i % 2 == 1
        bg  = _ROW_ALT if alt else _WHITE
        for ci, val in enumerate([label, tval, mval]):
            c           = ws.cell(row=row, column=ci + 1)
            c.value     = val
            c.font      = _f(9)
            c.fill      = _fill(bg)
            c.border    = _BORDER
            c.alignment = Alignment(horizontal="left" if ci == 0 else "right",
                                    vertical="center")
            if ci > 0 and val is not None:
                c.number_format = "0.0"
        ws.row_dimensions[row].height = 18
        row += 1

    chart_end_row = row - 1

    # Build the clustered column chart
    chart         = BarChart()
    chart.type    = "col"
    chart.grouping = "clustered"
    chart.title   = f"{ticker} vs Peer Median — Gross Margin / Op Margin / ROE (%)"
    chart.style   = 10
    chart.y_axis.title  = "%"
    chart.y_axis.numFmt = "0.0"
    chart.x_axis.title  = "Metric"
    chart.width   = 26
    chart.height  = 15

    data_ref = Reference(ws, min_col=2, max_col=3,
                         min_row=chart_data_row, max_row=chart_end_row)
    cats_ref = Reference(ws, min_col=1,
                         min_row=chart_data_row + 1, max_row=chart_end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    try:
        chart.series[0].graphicalProperties.solidFill = _NAVY
        chart.series[1].graphicalProperties.solidFill = "AAAAAA"
    except Exception:
        pass

    from openpyxl.chart.legend import Legend as _Legend
    leg          = _Legend()
    leg.position = "b"
    leg.overlay  = False
    chart.legend = leg

    ws.add_chart(chart, f"A{row + 1}")

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, w in zip("ABCDEFGHIJ", [24, 10, 12, 14, 15, 14, 11, 10, 12, 16]):
        ws.column_dimensions[col].width = w


# ── GS-style chart XML post-processor ────────────────────────────────────────
# openpyxl's chart API produces minimal XML with no colours, wrong axis positions,
# and no professional styling.  After wb.save() we open the xlsx as a zip, extract
# the dynamic cell-reference formulas from the generated chart XML, and replace the
# entire chart document with hand-crafted OOXML that matches Goldman Sachs style.

_NS_C = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"

_CHART_HDR = (
    "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\n"
    '<chartSpace xmlns="http://schemas.openxmlformats.org/drawingml/2006/chart"'
    ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
    ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
)
_CHART_FTR = "</chartSpace>"

_CHART_SPACE_SPR = (
    "<spPr>"
    "<a:solidFill><a:srgbClr val='FFFFFF'/></a:solidFill>"
    "<a:ln><a:noFill/></a:ln>"
    "<a:effectLst/>"
    "</spPr>"
)


def _txpr(sz: int = 900, color: str = "595959") -> str:
    return (
        f"<txPr><a:bodyPr/><a:lstStyle/>"
        f"<a:p><a:pPr><a:defRPr sz='{sz}' b='0'>"
        f"<a:solidFill><a:srgbClr val='{color}'/></a:solidFill>"
        f"<a:latin typeface='Calibri'/>"
        f"</a:defRPr></a:pPr></a:p></txPr>"
    )


def _extract_refs(xml: str) -> list[str]:
    return re.findall(r"<f>([^<]+)</f>", xml)


def _gs_price_chart_xml(xml_str: str, ticker: str) -> str:
    """Full GS-style line chart: navy stock line, grey-dashed S&P, professional axes."""
    refs    = _extract_refs(xml_str)
    ymin_m  = re.search(r'<min val="([^"]+)"', xml_str)
    ymax_m  = re.search(r'<max val="([^"]+)"', xml_str)
    skip_m  = re.search(r'tickLblSkip val="([^"]+)"', xml_str)

    s1_cat  = refs[1] if len(refs) > 1 else f"'Price Chart'!$A$2:$A$126"
    s1_val  = refs[2] if len(refs) > 2 else f"'Price Chart'!$B$2:$B$126"
    s2_val  = refs[5] if len(refs) > 5 else f"'Price Chart'!$C$2:$C$126"

    ymin    = float(ymin_m.group(1)) if ymin_m else 85.0
    ymax    = float(ymax_m.group(1)) if ymax_m else 120.0
    skip    = skip_m.group(1) if skip_m else "21"

    title   = f"{ticker} vs. S&amp;P 500  —  6-Month Indexed Performance (Base = 100)"
    s_sheet = s1_cat.split("!")[0]  # e.g. 'Price Chart'

    ser1 = (
        "<ser>"
        "<idx val='0'/><order val='0'/>"
        f"<tx><strRef><f>{s_sheet}!B1</f></strRef></tx>"
        "<spPr>"
        "<a:ln w='25400' cap='rnd'>"
        "<a:solidFill><a:srgbClr val='003366'/></a:solidFill>"
        "<a:prstDash val='solid'/>"
        "</a:ln>"
        "</spPr>"
        "<marker><symbol val='none'/></marker>"
        f"<cat><strRef><f>{s1_cat}</f></strRef></cat>"
        f"<val><numRef><f>{s1_val}</f></numRef></val>"
        "<smooth val='1'/>"
        "</ser>"
    )
    ser2 = (
        "<ser>"
        "<idx val='1'/><order val='1'/>"
        f"<tx><strRef><f>{s_sheet}!C1</f></strRef></tx>"
        "<spPr>"
        "<a:ln w='15875' cap='rnd'>"
        "<a:solidFill><a:srgbClr val='7F7F7F'/></a:solidFill>"
        "<a:prstDash val='sysDash'/>"
        "</a:ln>"
        "</spPr>"
        "<marker><symbol val='none'/></marker>"
        f"<cat><strRef><f>{s1_cat}</f></strRef></cat>"
        f"<val><numRef><f>{s2_val}</f></numRef></val>"
        "<smooth val='1'/>"
        "</ser>"
    )
    cat_ax = (
        "<catAx>"
        "<axId val='10'/>"
        "<scaling><orientation val='minMax'/></scaling>"
        "<delete val='0'/>"
        "<axPos val='b'/>"
        "<spPr><a:ln w='9525'>"
        "<a:solidFill><a:srgbClr val='D0D0D0'/></a:solidFill>"
        "</a:ln></spPr>"
        + _txpr(900, "595959") +
        "<majorTickMark val='none'/>"
        "<minorTickMark val='none'/>"
        "<tickLblPos val='low'/>"
        "<crossAx val='100'/>"
        f"<tickLblSkip val='{skip}'/>"
        "<noMultiLvlLbl val='1'/>"
        "</catAx>"
    )
    val_ax = (
        "<valAx>"
        "<axId val='100'/>"
        "<scaling>"
        "<orientation val='minMax'/>"
        f"<min val='{ymin:.2f}'/>"
        f"<max val='{ymax:.2f}'/>"
        "</scaling>"
        "<delete val='0'/>"
        "<axPos val='l'/>"
        "<majorGridlines>"
        "<spPr><a:ln w='9525' cmpd='sng'>"
        "<a:solidFill><a:srgbClr val='E8E8E8'/></a:solidFill>"
        "<a:prstDash val='solid'/>"
        "</a:ln></spPr>"
        "</majorGridlines>"
        "<spPr><a:ln><a:noFill/></a:ln></spPr>"
        + _txpr(900, "595959") +
        "<numFmt formatCode='0.0' sourceLinked='0'/>"
        "<majorTickMark val='none'/>"
        "<minorTickMark val='none'/>"
        "<crossAx val='10'/>"
        "</valAx>"
    )
    legend = (
        "<legend>"
        "<legendPos val='b'/>"
        "<overlay val='0'/>"
        "<spPr><a:noFill/><a:ln><a:noFill/></a:ln></spPr>"
        + _txpr(900, "595959") +
        "</legend>"
    )
    title_xml = (
        "<title>"
        "<tx><rich>"
        "<a:bodyPr/><a:lstStyle/>"
        "<a:p><a:pPr>"
        "<a:defRPr b='1' sz='1100'>"
        "<a:solidFill><a:srgbClr val='003366'/></a:solidFill>"
        "<a:latin typeface='Calibri'/>"
        "</a:defRPr></a:pPr>"
        "<a:r>"
        "<a:rPr lang='en-US' b='1' sz='1100' dirty='0'>"
        "<a:solidFill><a:srgbClr val='003366'/></a:solidFill>"
        "<a:latin typeface='Calibri'/>"
        "</a:rPr>"
        f"<a:t>{title}</a:t>"
        "</a:r></a:p>"
        "</rich></tx>"
        "<overlay val='0'/>"
        "<spPr><a:noFill/><a:ln><a:noFill/></a:ln></spPr>"
        "</title>"
    )
    plot_area_spr = (
        "<spPr>"
        "<a:solidFill><a:srgbClr val='FFFFFF'/></a:solidFill>"
        "<a:ln><a:noFill/></a:ln>"
        "<a:effectLst/>"
        "</spPr>"
    )
    return (
        _CHART_HDR
        + "<roundedCorners val='0'/>"
        + "<chart>"
        + title_xml
        + "<autoTitleDeleted val='0'/>"
        + "<plotArea>"
        + "<layout/>"
        + plot_area_spr
        + "<lineChart>"
        + "<grouping val='standard'/>"
        + "<varyColors val='0'/>"
        + ser1 + ser2
        + "<axId val='10'/><axId val='100'/>"
        + "</lineChart>"
        + cat_ax + val_ax
        + "</plotArea>"
        + legend
        + "<plotVisOnly val='1'/>"
        + "<dispBlanksAs val='gap'/>"
        + "</chart>"
        + _CHART_SPACE_SPR
        + _CHART_FTR
    )


def _gs_comp_chart_xml(xml_str: str) -> str:
    """Full GS-style clustered column chart: navy/steel bars with data labels."""
    refs   = _extract_refs(xml_str)

    s1_title = refs[0] if len(refs) > 0 else "'Competitive Analysis'!B19"
    s1_cat   = refs[1] if len(refs) > 1 else "'Competitive Analysis'!$A$20:$A$22"
    s1_val   = refs[2] if len(refs) > 2 else "'Competitive Analysis'!$B$20:$B$22"
    s2_title = refs[3] if len(refs) > 3 else "'Competitive Analysis'!C19"
    s2_cat   = refs[4] if len(refs) > 4 else s1_cat
    s2_val   = refs[5] if len(refs) > 5 else "'Competitive Analysis'!$C$20:$C$22"

    def _dlbls(fg: str) -> str:
        return (
            "<dLbls>"
            "<numFmt formatCode='0.0' sourceLinked='0'/>"
            "<spPr><a:noFill/><a:ln><a:noFill/></a:ln></spPr>"
            "<txPr><a:bodyPr/><a:lstStyle/>"
            "<a:p><a:pPr>"
            f"<a:defRPr sz='900' b='1'>"
            f"<a:solidFill><a:srgbClr val='{fg}'/></a:solidFill>"
            "<a:latin typeface='Calibri'/>"
            "</a:defRPr></a:pPr></a:p></txPr>"
            "<showLegendKey val='0'/>"
            "<showVal val='1'/>"
            "<showCatName val='0'/>"
            "<showSerName val='0'/>"
            "<showPercent val='0'/>"
            "<showBubbleSize val='0'/>"
            "<dLblPos val='outEnd'/>"
            "</dLbls>"
        )

    ser1 = (
        "<ser>"
        "<idx val='0'/><order val='0'/>"
        f"<tx><strRef><f>{s1_title}</f></strRef></tx>"
        "<spPr>"
        "<a:solidFill><a:srgbClr val='003366'/></a:solidFill>"
        "<a:ln><a:noFill/></a:ln>"
        "</spPr>"
        + _dlbls("003366")
        + f"<cat><strRef><f>{s1_cat}</f></strRef></cat>"
        + f"<val><numRef><f>{s1_val}</f></numRef></val>"
        + "</ser>"
    )
    ser2 = (
        "<ser>"
        "<idx val='1'/><order val='1'/>"
        f"<tx><strRef><f>{s2_title}</f></strRef></tx>"
        "<spPr>"
        "<a:solidFill><a:srgbClr val='8EB4E3'/></a:solidFill>"
        "<a:ln><a:noFill/></a:ln>"
        "</spPr>"
        + _dlbls("2D5F8A")
        + f"<cat><strRef><f>{s2_cat}</f></strRef></cat>"
        + f"<val><numRef><f>{s2_val}</f></numRef></val>"
        + "</ser>"
    )
    cat_ax = (
        "<catAx>"
        "<axId val='10'/>"
        "<scaling><orientation val='minMax'/></scaling>"
        "<delete val='0'/>"
        "<axPos val='b'/>"
        "<spPr><a:ln><a:noFill/></a:ln></spPr>"
        + _txpr(1000, "2C2C2C") +
        "<majorTickMark val='none'/>"
        "<minorTickMark val='none'/>"
        "<tickLblPos val='low'/>"
        "<crossAx val='100'/>"
        "<auto val='0'/>"
        "<noMultiLvlLbl val='1'/>"
        "</catAx>"
    )
    val_ax = (
        "<valAx>"
        "<axId val='100'/>"
        "<scaling><orientation val='minMax'/></scaling>"
        "<delete val='0'/>"
        "<axPos val='l'/>"
        "<majorGridlines>"
        "<spPr><a:ln w='9525'>"
        "<a:solidFill><a:srgbClr val='E8E8E8'/></a:solidFill>"
        "<a:prstDash val='solid'/>"
        "</a:ln></spPr>"
        "</majorGridlines>"
        "<spPr><a:ln><a:noFill/></a:ln></spPr>"
        + _txpr(900, "595959") +
        "<numFmt formatCode='0.0' sourceLinked='0'/>"
        "<majorTickMark val='none'/>"
        "<minorTickMark val='none'/>"
        "<crossAx val='10'/>"
        "</valAx>"
    )
    legend = (
        "<legend>"
        "<legendPos val='b'/>"
        "<overlay val='0'/>"
        "<spPr><a:noFill/><a:ln><a:noFill/></a:ln></spPr>"
        + _txpr(900, "595959") +
        "</legend>"
    )
    plot_area_spr = (
        "<spPr>"
        "<a:solidFill><a:srgbClr val='FFFFFF'/></a:solidFill>"
        "<a:ln><a:noFill/></a:ln>"
        "<a:effectLst/>"
        "</spPr>"
    )
    return (
        _CHART_HDR
        + "<roundedCorners val='0'/>"
        + "<chart>"
        + "<autoTitleDeleted val='1'/>"
        + "<plotArea>"
        + "<layout/>"
        + plot_area_spr
        + "<barChart>"
        + "<barDir val='col'/>"
        + "<grouping val='clustered'/>"
        + "<varyColors val='0'/>"
        + ser1 + ser2
        + "<gapWidth val='100'/>"
        + "<overlap val='0'/>"
        + "<axId val='10'/><axId val='100'/>"
        + "</barChart>"
        + cat_ax + val_ax
        + "</plotArea>"
        + legend
        + "<plotVisOnly val='1'/>"
        + "<dispBlanksAs val='gap'/>"
        + "</chart>"
        + _CHART_SPACE_SPR
        + _CHART_FTR
    )


def _polish_charts(output_path: str, ticker: str) -> None:
    """Replace openpyxl chart XML in the saved xlsx with GS-style professional OOXML."""
    tmp = output_path + ".~tmp"
    try:
        with zipfile.ZipFile(output_path, "r") as zin:
            with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    name = item.filename
                    if name.startswith("xl/charts/chart") and name.endswith(".xml"):
                        xml_str = data.decode("utf-8")
                        if "<lineChart>" in xml_str:
                            data = _gs_price_chart_xml(xml_str, ticker).encode("utf-8")
                        elif "<barChart>" in xml_str:
                            data = _gs_comp_chart_xml(xml_str).encode("utf-8")
                    zout.writestr(item, data)
        os.replace(tmp, output_path)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


def build_excel(ticker: str, stats: dict, fin_data: dict,
                price_history, sp500_history, markdown: str,
                news_sentiment: dict | None,
                dcf_result: dict | None,
                research: dict | None,
                comp_result: dict | None,
                output_path: str) -> None:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_snapshot_sheet(wb, ticker, stats)
    _build_chart_sheet(wb, ticker, price_history, sp500_history)
    _build_analysis_sheet(wb, markdown)
    _build_bull_bear_sheet(wb, markdown)
    _build_income_sheet(wb, fin_data)
    _build_balance_sheet_sheet(wb, fin_data)
    _build_cashflow_sheet(wb, fin_data)
    _build_news_sheet(wb, news_sentiment)
    _build_dcf_sheet(wb, dcf_result, ticker)
    _build_thesis_sheet(wb, research)
    _build_comps_sheet(wb, research)
    _build_earnings_sheet(wb, research)
    _build_competitive_sheet(wb, comp_result)
    wb.save(output_path)
    _polish_charts(output_path, ticker)
